# -*- coding: utf-8 -*-
"""
财报跟踪平台 · 数据管道
========================
读取锡/锌两个 Excel 数据源，抽取公司财报产量数据，生成 data/data.js
（window.SITE_DATA = {...}，前端用 <script src> 引入，file:// 协议可直接运行）。

扩展新品种的方法：
  1. 在下方 COMMODITIES 注册表中新增一个条目（key/名称/excel 路径/单位/抽取函数/日历/口径说明）；
  2. 编写对应的抽取函数（返回统一的数据结构，参考 extract_tin / extract_zinc）；
  3. 运行本脚本，前端会自动渲染新的标签页，无需改动前端代码。

运行：python build_site.py
"""

import json
import re
import datetime
from pathlib import Path

from openpyxl import load_workbook

BASE_DIR = Path(__file__).resolve().parent
OUT_FILE = BASE_DIR / "data" / "data.js"

BUILD_DATE = datetime.date.today()

# ---------------------------------------------------------------------------
# 披露日历常量（手工维护；status 由本脚本按构建日期自动判定：
#   日期已过 -> "已披露待核"，未来 -> "待披露"）
# 字段：date(YYYY-MM-DD), approx(是否约估日期), company, event(事件说明)
# commodity 由注册表条目挂接时自动补上。
# ---------------------------------------------------------------------------
TIN_CALENDAR = [
    # Alphamin：季报≈1/4/7/10 月中产量快报 + 月底完整财报
    {"date": "2026-10-15", "approx": True, "company": "Alphamin", "event": "2026Q3 产量快报"},
    {"date": "2026-10-30", "approx": True, "company": "Alphamin", "event": "2026Q3 完整财报"},
    # Metals X：季报≈1/4/7/10 月底
    {"date": "2026-10-30", "approx": True, "company": "Metals X", "event": "2026Q3 季度报告"},
    # PT Timah：4 月底年报+Q1、7 月底半年报、11 月初 9M
    {"date": "2026-11-05", "approx": True, "company": "PT Timah", "event": "2026 年 9M（前三季）报告"},
    # 明苏尔（Minsur）：≈3 月初/5 月中/8 月中/11 月中
    {"date": "2026-08-15", "approx": True, "company": "明苏尔", "event": "2026Q2 季报"},
    {"date": "2026-11-15", "approx": True, "company": "明苏尔", "event": "2026Q3 季报"},
    # MSC：≈2 月/5 月/8 月/11 月下旬（季报不披露吨数）
    {"date": "2026-08-28", "approx": True, "company": "MSC", "event": "2026Q2 季报（无吨数）"},
    {"date": "2026-11-27", "approx": True, "company": "MSC", "event": "2026Q3 季报（无吨数）"},
    # 云锡/兴业/华锡（A 股）：4 月下旬年报+一季报、8 月下旬中报、10 月下旬三季报
    {"date": "2026-08-28", "approx": True, "company": "云南锡业", "event": "2026 中报"},
    {"date": "2026-10-28", "approx": True, "company": "云南锡业", "event": "2026 三季报"},
    {"date": "2026-08-28", "approx": True, "company": "兴业银锡", "event": "2026 中报"},
    {"date": "2026-10-28", "approx": True, "company": "兴业银锡", "event": "2026 三季报"},
    {"date": "2026-08-28", "approx": True, "company": "华锡有色", "event": "2026 中报"},
    {"date": "2026-10-28", "approx": True, "company": "华锡有色", "event": "2026 三季报"},
]

ZINC_CALENDAR = [
    # 2026Q2 财报季（已知/约估日期；季报≈季后 3-4 周）
    {"date": "2026-07-20", "approx": False, "company": "South32", "event": "FY26 Q4（6 月季）产量报告"},
    {"date": "2026-07-21", "approx": False, "company": "Boliden", "event": "2026Q2 财报"},
    {"date": "2026-07-21", "approx": True, "company": "MMG", "event": "2026Q2 产量报告"},
    {"date": "2026-07-23", "approx": False, "company": "Teck", "event": "2026Q2 财报"},
    {"date": "2026-07-23", "approx": False, "company": "Newmont", "event": "2026Q2 财报"},
    {"date": "2026-07-29", "approx": False, "company": "Glencore", "event": "2026 H1 产量报告"},
    {"date": "2026-07-29", "approx": False, "company": "Ivanhoe", "event": "2026Q2 财报（Kipushi）"},
    {"date": "2026-08-03", "approx": True, "company": "Peñoles", "event": "2026Q2 财报"},
    {"date": "2026-08-05", "approx": False, "company": "Nexa", "event": "2026Q2 财报"},
    {"date": "2026-08-06", "approx": True, "company": "Korea Zinc", "event": "2026Q2 财报"},
    {"date": "2026-08-07", "approx": True, "company": "Mitsui Mining", "event": "2027/3 期 Q1 财报"},
    {"date": "2026-10-03", "approx": True, "company": "Vedanta", "event": "Q2FY27 产量公告"},
    # 注：Nyrstar 不披露季度产量，不列入日历。
]

# ---------------------------------------------------------------------------
# 锡 · 更新日志常量（锡 Excel 无更新日志 sheet，在此手工维护）
# ---------------------------------------------------------------------------
TIN_CHANGELOG = [
    {"date": "2026-07-26", "content": "锡表更新至 Alphamin 26Q2 / 其余 26Q1，明细见 Excel", "source": "海外主要公司产量.xlsx"},
    {"date": "2026-07-26", "content": "接入「季度补充」sheet：云锡（产品锡/锡矿拟合）、华锡（锡精矿/锡锭）、兴业（矿产锡）、明苏尔（矿产锡SR+B2/精炼锡Pisco）季度序列；明苏尔 2025Q2 修正为 8390（原 8205 系 2024Q2 误填）、2025 全年 33820、2026Q1 7993；明苏尔精炼锡 2024 修正为 30926；明苏尔/Pisco 季度数据唯一来源改为季度补充表，删除脚本内补充常量。", "source": "海外主要公司产量.xlsx · 季度补充"},
]

# 注：明苏尔（Minsur）矿产锡季度序列与 Pisco 精炼锡序列原先以脚本常量维护，
# 现已统一改由 Excel「季度补充」sheet 读取（唯一来源），此处不再保留常量，避免双口径。

# ---------------------------------------------------------------------------
# 锡 · FY2026 指引常量（Sheet1 T 列为 2025 指引，2026 指引在此维护；
# 来源：PT Timah 官网 2026 年产量目标）
# ---------------------------------------------------------------------------
TIN_FY2026_GUIDE = {
    "PT Timah": "锡矿 30,000 吨（2026 年产量目标，官网）",
}

# 指引年化排除名单（季节性或状态特殊，不做机械年化）：
# Metro Q1 雨季、Nornickel Q4 冲量、South32（Cerro Matoso 已出售）
# 支持「公司·项目」精确写法：BOLIDEN·Garpenberg=2026 地震事件年（Q2 停产，H2 低产能爬坡，×2 年化失真）
# BOLIDEN·Tara=2026 指引下调年（Q2 磨矿指引 1.8→1.6Mt，开拓滞后意味着 H2 偏弱，×2 年化高估）
# 住友金属矿山=财年止 3 月（日历 Q1 年化×4 与 FY 错位）；Sigma/紫金锂/Ozernaya=爬坡期（×2/×4 失真）
NO_ANNUALIZE_GUIDE = {"Metro Mining", "Nornickel", "South32", "BOLIDEN·Garpenberg", "BOLIDEN·Tara",
                      "住友金属矿山", "Sigma Lithium", "紫金矿业·3Q+拉果错+湘源(LCE)", "Polymetal·Ozernaya Minin"}

# ---------------------------------------------------------------------------
# 锡 · 成本补充常量（官方财报 / MD&A / 券商模型，2026-07 核实）
# ---------------------------------------------------------------------------
# Alphamin AISC 序列中属于「指引值」（非实际）的季度：26Q2=19,043 为公司指引
ALPHAMIN_AISC_GUIDE_PERIODS = {"2026Q2"}
# Metals X 澳元成本折算美元的假设汇率（图表与 tooltip 注明）
AUDUSD_RATE = 0.65
MINSUR_COST_TEXT = "精炼锡现金成本 $7,751/t（1Q26，+7%）"
TIMAH_COST_TEXT = "券商估算 $21,686/t（2023 锚点，BRIDS 模型，估算）"
# 国内公司成本/毛利锚点（2025 年报）
TIN_DOMESTIC_COST = {
    ("refined", "云南锡业"): "锡锭毛利率 11.45%（2025 年报）",
    ("refined", "华锡有色"): "锡锭单位成本 +26.44%、毛利率 -6.77pct（2025 年报）",
    ("mine", "兴业银锡"): "整体毛利率 57.59%（-5.36pct，2025 年报）",
}

# ---------------------------------------------------------------------------
# 口径说明（展示在各品种页底部折叠区）
# ---------------------------------------------------------------------------
TIN_CALIBER_NOTES = [
    "产量/销量单位均为吨；矿表口径为矿产锡（精矿含锡），锭表为精炼锡。",
    "Metals X 为矿山 100% 口径（Renison 矿），非 50% 权益口径。",
    "PT Timah 单季值为从公司累计披露推算，非直接披露值。",
    "MSC 季报不披露吨数，仅有年度实际产量（Q4 行）。",
    "兴业银锡「矿产锡」= 锡金属（合格锡精矿含锡）+ 锡次金属（低品位锡精粉），两者并列相加、非包含关系；2021-2023 已修正为全口径。",
    "云锡/华锡/兴业/明苏尔的季度序列来自 Excel「季度补充」sheet（2023Q1 起）；明苏尔矿产锡 = San Rafael + B2 尾矿厂合计，明苏尔精炼锡 = Pisco 冶炼厂（Pisco 行不再单独重复展示）。",
    "带 † 标记的序列含估算/推算成分（如年度×季节分布拟合、累计相减、指引估算），备注原文见表格公司名悬停。",
    "云锡 2024 年存在两个口径：Sheet1 锭表 83478 = 锡锭 78454 + 锡材 5024；官方「产品锡」口径 2024 = 84800（季度补充表序列口径），两者并存。",
    "Alphamin AISC 单位为 US$/t；Metals X C1/AISC 单位为 A$/t（澳元）。",
]

ZINC_CALIBER_NOTES = [
    "锌矿产量单位为万金属吨；锌锭冶炼产量单位为万吨精炼锌。",
    "Nyrstar 不披露季度产量财报，冶炼合计中不含 Nyrstar。",
    "锌矿总计行为表内公司合计；OZ 矿（Ozernoye）为非财报披露口径。",
    "资本开支币种各异，未折算：Teck 部分口径为加元（CAD）、Boliden 为瑞典克朗（SEK）、Vedanta/HZL 为印度卢比（INR crore），比较时注意口径说明列。",
    "Ivanhoe Kipushi 产量与资本开支指引为项目 100% 口径。",
]

# ---------------------------------------------------------------------------
# 铝 · 披露日历（公司粒度；季报≈季后数周，财年口径见口径说明）
# ---------------------------------------------------------------------------
ALUMINUM_CALENDAR = [
    # 美铝：≈季后 2-3 周
    {"date": "2026-07-15", "approx": True, "company": "美铝 Alcoa", "event": "2026Q2 财报"},
    {"date": "2026-10-21", "approx": True, "company": "美铝 Alcoa", "event": "2026Q3 财报"},
    # 力拓：≈季后 2 周（产量报告）
    {"date": "2026-07-15", "approx": True, "company": "力拓 Rio Tinto", "event": "2026Q2 产量报告"},
    {"date": "2026-10-14", "approx": True, "company": "力拓 Rio Tinto", "event": "2026Q3 产量报告"},
    # 海德鲁：≈季后 4 周
    {"date": "2026-07-24", "approx": True, "company": "海德鲁 Norsk Hydro", "event": "2026Q2 财报"},
    {"date": "2026-10-28", "approx": True, "company": "海德鲁 Norsk Hydro", "event": "2026Q3 财报"},
    # South32：≈1/4/7/10 月 20 日（财年止 6 月）
    {"date": "2026-07-20", "approx": False, "company": "South32", "event": "FY26 Q4（6月季）产量报告"},
    {"date": "2026-10-20", "approx": True, "company": "South32", "event": "FY27 Q1（9月季）产量报告"},
    # 印度三家：财年止 3 月，≈1/4/7/10 月底-次月初
    {"date": "2026-10-03", "approx": True, "company": "Vedanta", "event": "Q2FY27 产量公告"},
    {"date": "2026-08-15", "approx": True, "company": "Hindalco", "event": "Q1FY27 财报（预计 8 月中）"},
    {"date": "2026-08-15", "approx": True, "company": "NALCO", "event": "Q1FY27 财报"},
    # 俄铝：≈2 月/8 月（仅半年报/年报）
    {"date": "2026-08-28", "approx": True, "company": "俄铝 RUSAL", "event": "2026 半年报"},
    # Century：≈季后 3 周；Metro：≈季后 3-4 周（财年止 6 月）；Press Metal：≈2/5/8/11 月底
    {"date": "2026-07-22", "approx": True, "company": "Century Aluminum", "event": "2026Q2 财报（数据待补）"},
    {"date": "2026-07-27", "approx": True, "company": "Metro Mining", "event": "FY26 Q4 季报（数据待补）"},
    {"date": "2026-08-28", "approx": True, "company": "Press Metal", "event": "2026Q2 季报（数据待补）"},
    # 国内：≈4 月底/8 月底/10 月底
    {"date": "2026-08-31", "approx": True, "company": "中国铝业", "event": "2026 中报"},
    {"date": "2026-08-31", "approx": True, "company": "中国宏桥/宏桥控股", "event": "2026 中报"},
    {"date": "2026-08-31", "approx": True, "company": "云铝股份", "event": "2026 中报"},
    {"date": "2026-08-31", "approx": True, "company": "天山铝业", "event": "2026 中报"},
    {"date": "2026-08-31", "approx": True, "company": "神火股份", "event": "2026 中报"},
    {"date": "2026-10-30", "approx": True, "company": "中国铝业", "event": "2026 三季报"},
]

ALUMINUM_CALIBER_NOTES = [
    "产量单位均为万吨；铝土矿/氧化铝/电解铝三板块。",
    "权益口径：力拓三板块均为权益产量；海德鲁电解铝 = 挪威等 100% + Qatalum 50% 权益计入；South32 按各资产权益（Worsley 86% / Mozal 63.7% / 巴西 36%-40%）。",
    "Metro Mining 为湿吨口径；力拓铝土矿亦为湿吨。",
    "印度公司（Vedanta/Hindalco/NALCO）财年止 3 月（Q1FY = 自然季 Q2），NALCO 季度绝对量不披露、仅有财年值。",
    "俄铝 2022 年起停发季报，仅半年度/年度披露；多数 A 股公司（宏桥/云铝/天山/神火/创新等）亦为年度/半年度披露，季度列为 '-' 属正常，建议用年度视图。",
    "中国宏桥（1378.HK）与宏桥控股（002379.SZ）为同一资产，只列一行，不重复计列。",
    "总计行为口径混杂（权益/100% 合并/湿吨/财年值）直接求和，仅作量级参考。",
    "美铝 2026Q2 宣布收购 South32 铝资产（Worsley 等），后续留意并表时点。",
]

# ---------------------------------------------------------------------------
# 镍 · 披露日历（公司粒度；主波 1/4/7/10 月底；South32 已退出镍业务不再列）
# ---------------------------------------------------------------------------
NICKEL_CALENDAR = [
    # 2026Q2 财报季
    {"date": "2026-07-27", "approx": True, "company": "Nornickel", "event": "2026Q2/H1 产量报告（约 7/27 当周）"},
    {"date": "2026-07-29", "approx": False, "company": "Glencore", "event": "2026 H1 产量报告"},
    {"date": "2026-07-29", "approx": False, "company": "Eramet", "event": "2026 半年报"},
    {"date": "2026-07-29", "approx": False, "company": "Nickel Industries", "event": "2026Q2 季报（已定）"},
    {"date": "2026-07-31", "approx": True, "company": "Antam", "event": "2026Q2 季报（约 7 月底）"},
    {"date": "2026-08-03", "approx": True, "company": "Merdeka(MBMA)", "event": "2026Q2 季报（约 7 月底-8 月初）"},
    {"date": "2026-08-05", "approx": True, "company": "Ambatovy", "event": "2026Q2 更新（约 8 月初，住友商事/AMRI）"},
    {"date": "2026-08-07", "approx": True, "company": "住友金属矿山", "event": "FY26 Q1 财报（约 8 月上旬）"},
    {"date": "2026-08-12", "approx": False, "company": "Sherritt", "event": "2026Q2 财报"},
    # 印尼中资：中报 8 月中下旬
    {"date": "2026-08-25", "approx": True, "company": "华友钴业", "event": "2026 中报（8 月中下旬）"},
    {"date": "2026-08-25", "approx": True, "company": "中伟股份", "event": "2026 中报（8 月中下旬）"},
    {"date": "2026-08-25", "approx": True, "company": "格林美", "event": "2026 中报（8 月中下旬）"},
    {"date": "2026-08-25", "approx": True, "company": "力勤资源", "event": "2026 中报（约 8 月下旬）"},
    # 往后按季循环（1/4/7/10 月底为主波）
    {"date": "2026-10-26", "approx": True, "company": "Anglo American", "event": "2026Q3 产量报告（约 10 月下旬）"},
    {"date": "2026-10-27", "approx": True, "company": "Vale", "event": "2026Q3 产量报告（约 10 月下旬）"},
    {"date": "2026-10-28", "approx": True, "company": "Nornickel", "event": "2026Q3/9M 产量报告（约 10 月底）"},
    {"date": "2026-10-29", "approx": True, "company": "Eramet", "event": "2026Q3 销售/产量更新"},
    {"date": "2026-10-29", "approx": True, "company": "Nickel Industries", "event": "2026Q3 季报"},
    {"date": "2026-10-30", "approx": True, "company": "Glencore", "event": "2026 9M 产量报告"},
]

NICKEL_CALIBER_NOTES = [
    "产量单位均为吨镍（金属量）；一级镍 / 镍铁NPI / 印尼中间品三板块，印尼中间品为边际重点板块（排第一位）。",
    "PTVI（Vale 印尼）自 3Q24 起不再 100% 并表，按 80% offtake 入表；100% 口径 ≈ offtake ÷ 0.8（推算）。",
    "Nickel Industries 的 RKEF / HNC 为 100% 项目基准（权益分别为 80% / 10%）；Eramet Weda Bay 同为 100% 口径（权益 38.7%）。",
    "华友钴业为 100% 项目口径（权益约 13-16.5 万吨，券商估）；中伟/格林美为券商测算或含参股出货口径，与官方披露口径不同。",
    "状态跟踪行：Ambatovy 待售 AMRI（住友 5/1 签约零对价出售 54.17%）、Anglo 镍待售 MMG（卡在欧盟反垄断审批）、South32 镍已售 CoreX（2025/12 交割，退出镍业务）、Sandouville/Koniambo 关停、BHP Nickel West 停产（2027/2 前复审）、Sherritt 因古巴制裁停摆。",
    "总计行为口径混杂（权益/100%/offtake/券商测算）直接求和，仅作量级参考。",
]

# ---------------------------------------------------------------------------
# 铜 · 披露日历（公司粒度；季后约 3 周为主波，KGHM 每月 22-25 日月度产销）
# ---------------------------------------------------------------------------
COPPER_CALENDAR = [
    {"date": "2026-07-27", "approx": True, "company": "Nornickel", "event": "2026 H1 产量报告（约 7/27 当周）"},
    {"date": "2026-07-28", "approx": False, "company": "First Quantum", "event": "2026Q2 财报（盘后）"},
    {"date": "2026-07-29", "approx": False, "company": "Glencore", "event": "2026 H1 产量报告"},
    {"date": "2026-08-11", "approx": False, "company": "MMG", "event": "2026 中报"},
    {"date": "2026-08-20", "approx": True, "company": "KGHM", "event": "2026 H1 报告（约 8 月中下旬；每月 22-25 日月度产销）"},
    {"date": "2026-08-25", "approx": True, "company": "紫金矿业", "event": "2026 中报（8 月下旬）"},
    {"date": "2026-08-25", "approx": True, "company": "洛阳钼业", "event": "2026 中报（8 月下旬）"},
    {"date": "2026-08-25", "approx": True, "company": "江西铜业", "event": "2026 中报（8 月下旬）"},
    {"date": "2026-08-25", "approx": True, "company": "铜陵有色", "event": "2026 中报（8 月下旬）"},
    {"date": "2026-08-25", "approx": True, "company": "云南铜业", "event": "2026 中报（8 月下旬）"},
    {"date": "2026-10-15", "approx": False, "company": "MMG", "event": "2026Q3 产量报告"},
    # 季后约 3 周主波（1/4/7/10 月下旬）
    {"date": "2026-10-22", "approx": True, "company": "FreePort", "event": "2026Q3 财报"},
    {"date": "2026-10-22", "approx": True, "company": "BHP", "event": "FY27 Q1 运营回顾"},
    {"date": "2026-10-22", "approx": True, "company": "Vale", "event": "2026Q3 产量报告"},
    {"date": "2026-10-22", "approx": True, "company": "Teck", "event": "2026Q3 财报"},
    {"date": "2026-10-23", "approx": True, "company": "Anglo American", "event": "2026Q3 产量报告"},
    {"date": "2026-10-23", "approx": True, "company": "Antofagasta", "event": "2026Q3 产量报告"},
    {"date": "2026-10-23", "approx": True, "company": "SouthernCopper", "event": "2026Q3 财报"},
    {"date": "2026-10-28", "approx": True, "company": "First Quantum", "event": "2026Q3 财报"},
]

COPPER_CALIBER_NOTES = [
    "铜矿板块单位为 kt（千吨），国内铜企板块单位为万吨，两板块单位不同、总计行各自独立。",
    "FreePort 为并表可回收磅口径（1M lbs = 0.4536kt，PTFI 持股 48.76% 但 100% 并表）；Rio Tinto 为 consolidated 口径（Kennecott/OT 100% + Escondida 30% 权益）。",
    "BHP Escondida 按 100% 计入（权益 57.5%），财年止 6 月，FY2026 指引列为 FY27 财年口径（1650-1800kt）。",
    "South32 Sierra Gorda 为 45% 应付铜（KGHM 持 55%，同一资产两处出现）；Teck QB 按 100%（持股 60%）。",
    "Anglo-Teck 合并仅剩中国审批（2026/9-2027/3 窗口），交割后需口径切换。",
    "Glencore 25Q3/Q4 为推算（FY851.6 - 9M583.5）；Nornickel 分本部/外贝加尔（Bystrinsky）两行；MMG 为 100% 基准。",
    "国内板块（紫金/洛钼/MMG/江铜/铜陵/云铜）单位万吨，多数仅年报/中报披露产量，季度列为空属正常。",
    "总计行为口径混杂（权益/并表/100%/财年映射）直接求和，仅作量级参考。",
]

# ---------------------------------------------------------------------------
# 锂 · 披露日历（澳矿季报 1/4/7/10 月末为主波；国内中报 8 月下旬、年报 3-4 月）
# ---------------------------------------------------------------------------
LITHIUM_CALENDAR = [
    # 2026Q2（澳矿为 FY26 Q4）财报季
    {"date": "2026-07-30", "approx": True, "company": "Pilbara Minerals", "event": "FY26 Q4（2026Q2）季报（7 月下旬）"},
    {"date": "2026-07-30", "approx": True, "company": "Mineral Resources", "event": "FY26 Q4（2026Q2）季报（7 月下旬）"},
    {"date": "2026-07-31", "approx": True, "company": "IGO", "event": "FY26 Q4（2026Q2）季报（7 月底）"},
    {"date": "2026-07-29", "approx": True, "company": "Liontown", "event": "FY26 Q4（2026Q2）季报（7 月下旬）"},
    {"date": "2026-07-30", "approx": True, "company": "Core Lithium", "event": "FY26 Q4（2026Q2）季报（7 月下旬）"},
    {"date": "2026-08-14", "approx": True, "company": "Sigma Lithium", "event": "2026Q2 季报（8 月中）"},
    {"date": "2026-08-20", "approx": True, "company": "SQM", "event": "2026Q2 财报（8 月中下旬）"},
    {"date": "2026-08-25", "approx": True, "company": "赣锋锂业", "event": "2026 中报（8 月下旬）"},
    {"date": "2026-08-25", "approx": True, "company": "中矿资源", "event": "2026 中报（8 月下旬）"},
    {"date": "2026-08-25", "approx": True, "company": "华友钴业", "event": "2026 中报（8 月下旬）"},
    {"date": "2026-08-25", "approx": True, "company": "盛新锂能", "event": "2026 中报（8 月下旬）"},
    {"date": "2026-08-25", "approx": True, "company": "雅化集团", "event": "2026 中报（8 月下旬）"},
    {"date": "2026-08-25", "approx": True, "company": "天齐锂业", "event": "2026 中报（8 月下旬）"},
    {"date": "2026-08-25", "approx": True, "company": "天华新能", "event": "2026 中报（8 月下旬）"},
    {"date": "2026-08-25", "approx": True, "company": "紫金矿业", "event": "2026 中报（8 月下旬）"},
    {"date": "2026-08-25", "approx": True, "company": "盐湖股份", "event": "2026 中报（8 月下旬）"},
    {"date": "2026-08-25", "approx": True, "company": "藏格矿业", "event": "2026 中报（8 月下旬）"},
    {"date": "2026-08-25", "approx": True, "company": "西藏矿业", "event": "2026 中报（8 月下旬）"},
    {"date": "2026-08-25", "approx": True, "company": "融捷股份", "event": "2026 中报（8 月下旬）"},
    {"date": "2026-08-25", "approx": True, "company": "永兴材料", "event": "2026 中报（8 月下旬）"},
    {"date": "2026-08-25", "approx": True, "company": "江特电机", "event": "2026 中报（8 月下旬）"},
    # 往后按季循环
    {"date": "2026-10-16", "approx": True, "company": "Rio Tinto", "event": "2026Q3 产量报告（10 月中）"},
    {"date": "2026-10-28", "approx": True, "company": "Pilbara Minerals", "event": "FY27 Q1（2026Q3）季报（10 月下旬）"},
    {"date": "2026-10-29", "approx": True, "company": "Mineral Resources", "event": "FY27 Q1（2026Q3）季报（10 月下旬）"},
    {"date": "2026-10-30", "approx": True, "company": "IGO", "event": "FY27 Q1（2026Q3）季报（10 月底）"},
]

LITHIUM_CALIBER_NOTES = [
    "锂资源板块单位为吨 LCE：锂辉石精矿按 SC6×0.125 折 LCE（精矿原值在各公司备注注明）；盐湖/云母为碳酸锂当量直接值。",
    "锂盐冶炼板块单位为吨锂盐实物量（碳酸锂+氢氧化锂合计），与资源端 LCE 口径不同，两板块总计行各自独立。",
    "澳矿（Pilbara/MinRes/IGO/Liontown/Core）财年止 6 月，已按日历季度映射（FY Q2=日历年 Q4 等）。",
    "Rio Tinto 2025/3 完成收购 Arcadium（2024/1 由 Allkem+Livent 合并），2023-2024 为 Arcadium 回溯口径（Olaroz/Fenix/Mt Cattlin 等），2025 起为 Rio 锂板块披露口径。",
    "SQM 为锂盐销量口径（产量不单独披露）；Greenbushes 按 100% 口径（IGO 24.99%/天齐 51%/雅保 24.01%）。",
    "中企（赣锋/中矿/华友/盛新/雅化/紫金/盐湖/藏格/西藏矿业/融捷/永兴/江特）多数仅年报/中报披露，季度列为平台拟合（斜体）。",
    "总计行为口径混杂（权益/100%/销量/拟合）直接求和，仅作量级参考。",
]

# ---------------------------------------------------------------------------
# COMMODITIES 注册表：加新品种 = 加一个条目 + 一个抽取函数
# ---------------------------------------------------------------------------
COMMODITIES = {
    "tin": {
        "name": "锡",
        "excel": r"D:\拷贝文件\E\永安\周报数据更新\进出口库存\海外主要公司产量.xlsx",
        "unit_mine": "吨（精矿含锡）",
        "unit_refined": "吨（精炼锡）",
        "extract": "extract_tin",
        "calendar": TIN_CALENDAR,
        "caliber_notes": TIN_CALIBER_NOTES,
        "default_view": "quarter",   # 品种页默认视图：quarter=季度 / year=年度
    },
    "zinc": {
        "name": "锌",
        "excel": r"D:\拷贝文件\E\永安\锌\全球锌企季度产量梳理.xlsx",
        "unit_mine": "万金属吨",
        "unit_refined": "万吨精炼锌",
        "extract": "extract_zinc",
        "calendar": ZINC_CALENDAR,
        "caliber_notes": ZINC_CALIBER_NOTES,
        "default_view": "quarter",
    },
    "aluminum": {
        "name": "铝",
        "excel": r"D:\拷贝文件\E\永安\铝\全球铝企季度产量梳理.xlsx",
        "extract": "extract_aluminum",
        "calendar": ALUMINUM_CALENDAR,
        "caliber_notes": ALUMINUM_CALIBER_NOTES,
        "default_view": "year",      # 铝年度/半年度披露公司居多，默认年度视图
    },
    "nickel": {
        "name": "镍",
        "excel": r"D:\拷贝文件\E\永安\镍\全球镍企季度产量梳理.xlsx",
        "extract": "extract_nickel",
        "calendar": NICKEL_CALENDAR,
        "caliber_notes": NICKEL_CALIBER_NOTES,
        "default_view": "quarter",
    },
    "copper": {
        "name": "铜",
        "excel": r"D:\拷贝文件\E\永安\铜\全球铜企季度产量梳理.xlsx",
        "extract": "extract_copper",
        "calendar": COPPER_CALENDAR,
        "caliber_notes": COPPER_CALIBER_NOTES,
        "default_view": "quarter",
    },
    "lithium": {
        "name": "锂",
        "excel": r"D:\拷贝文件\E\永安\锂\全球锂企季度产量梳理.xlsx",
        "extract": "extract_lithium",
        "calendar": LITHIUM_CALENDAR,
        "caliber_notes": LITHIUM_CALIBER_NOTES,
        "default_view": "quarter",
    },
}


# ---------------------------------------------------------------------------
# 工具函数
# ---------------------------------------------------------------------------
def num(v):
    """把单元格原始值转成 float 或 None。'/'、'-'、'n.a.'、None、公式（非纯数值算式）-> None。"""
    if v is None:
        return None
    if isinstance(v, (int, float)):
        return float(v)
    s = str(v).strip()
    if s in ("", "/", "-", "—", "n.a.", "N/A", "na"):
        return None
    if s.startswith("="):
        expr = s[1:]
        if re.fullmatch(r"[0-9.+\-*/() ]+", expr):
            try:
                return float(eval(expr, {"__builtins__": {}}, {}))
            except Exception:
                return None
        return None
    try:
        return float(s.replace(",", ""))
    except ValueError:
        return None


def txt(v):
    if v is None:
        return None
    s = str(v).strip()
    return s if s and s not in ("/", "-") else None


def quarter_sort_key(q):
    m = re.fullmatch(r"(\d{4})Q([1-4])", q)
    return (int(m.group(1)), int(m.group(2)))


def _short_q(q):
    """2026Q1 -> 26Q1（展示用短标签）。"""
    m = re.fullmatch(r"(\d{4})Q([1-4])", q or "")
    return f"{m.group(1)[2:]}Q{m.group(2)}" if m else (q or "")


def prior_year_period(p):
    """2026Q2 -> 2025Q2；'2025' -> '2024'；无法映射返回 None。"""
    m = re.fullmatch(r"(\d{4})Q([1-4])", p)
    if m:
        return f"{int(m.group(1)) - 1}Q{m.group(2)}"
    if re.fullmatch(r"\d{4}", p):
        return str(int(p) - 1)
    return None


def compute_yoy(data, provided=None):
    """对 data={period: value} 计算同比={period: float}。provided 为 Excel 已给的同比，作缺省回填。"""
    yoy = {}
    for p, v in data.items():
        if v is None:
            continue
        pp = prior_year_period(p)
        pv = data.get(pp)
        if pv:
            yoy[p] = v / pv - 1.0
        elif provided and p in provided and provided[p] is not None:
            yoy[p] = provided[p]
    return yoy


def sum_total(companies, periods):
    """各期合计（仅加总有值的公司），同比按「两期均有值的公司集合」计算，口径一致。"""
    total = {}
    for p in periods:
        vals = [c["data"].get(p) for c in companies]
        vals = [v for v in vals if v is not None]
        total[p] = round(sum(vals), 4) if vals else None
    yoy = {}
    for p in periods:
        pp = prior_year_period(p)
        if not pp:
            continue
        cur = sum(c["data"][p] for c in companies if c["data"].get(p) is not None and c["data"].get(pp) is not None)
        prv = sum(c["data"][pp] for c in companies if c["data"].get(p) is not None and c["data"].get(pp) is not None)
        if prv:
            yoy[p] = cur / prv - 1.0
    return {"data": total, "yoy": yoy}


def make_calendar(entries, commodity_name):
    out = []
    for e in entries:
        d = datetime.date.fromisoformat(e["date"])
        out.append({
            "date": e["date"],
            "approx": bool(e.get("approx")),
            "company": e["company"],
            "commodity": commodity_name,
            "event": e["event"],
            "status": "已披露待核" if d <= BUILD_DATE else "待披露",
        })
    return out


def section_stats(companies, quarters, title="矿山"):
    """总览卡片用：最新季度的披露进度与合计同比。title=首个板块简称（总览行标签）。"""
    qs = sorted(quarters, key=quarter_sort_key)
    latest = None
    for q in reversed(qs):
        if any(c["data"].get(q) is not None for c in companies):
            latest = q
            break
    if latest is None:
        return None
    disclosed = sum(1 for c in companies if c["data"].get(latest) is not None)
    pp = prior_year_period(latest)
    cur = prv = 0.0
    for c in companies:
        a, b = c["data"].get(latest), c["data"].get(pp)
        if a is not None and b is not None:
            cur += a
            prv += b
    return {
        "section_title": title,
        "latest_period": latest,
        "disclosed": disclosed,
        "total_companies": len(companies),
        "mine_total": round(sum(c["data"][latest] for c in companies if c["data"].get(latest) is not None), 4),
        "mine_yoy": (cur / prv - 1.0) if prv else None,
    }


# ---------------------------------------------------------------------------
# 锡抽取
# ---------------------------------------------------------------------------
def extract_tin(path):
    wb = load_workbook(path, data_only=False)
    s1 = wb["Sheet1"]

    # --- Sheet1 矿表（行 2-8）：J=公司,K=国家,L~O=2021~2024,Q=2025Q1,R=2025Q2,
    #     T=25指引,U=增产/减产原因,V=2025全年,W=2026Q1 ---
    mine_annual = {}   # name -> {period: val}
    mine_meta = {}     # name -> {country, guide, reason, note}
    for r in range(2, 9):
        name = txt(s1.cell(row=r, column=10).value)
        if not name:
            continue
        data = {}
        for i, year in enumerate(["2021", "2022", "2023", "2024"]):
            v = num(s1.cell(row=r, column=12 + i).value)
            if v is not None:
                data[year] = v
        v = num(s1.cell(row=r, column=22).value)   # V=2025全年
        if v is not None:
            data["2025"] = v
        for col, q in ((17, "2025Q1"), (18, "2025Q2"), (23, "2026Q1")):  # Q,R,W
            v = num(s1.cell(row=r, column=col).value)
            if v is not None:
                data[q] = v
        mine_annual[name] = data
        reason = txt(s1.cell(row=r, column=21).value)  # U
        note = None
        if name == "兴业银锡":
            note = "口径：矿产锡 = 锡金属(合格锡精矿含锡) + 锡次金属(低品位锡精粉)，并列相加、非包含"
        mine_meta[name] = {
            "country": txt(s1.cell(row=r, column=11).value),
            "guide": txt(s1.cell(row=r, column=20).value),  # T=25指引
            "reason": reason,
            "note": note,
        }

    # --- 季度序列：各公司 sheet ---
    quarterly = {n: {} for n in mine_annual}
    # Alphamin：行12-33，A=季度，C=AISC(US$/t)，D=产量(吨)，H=销量(吨)
    ws = wb["Alphamin"]
    alphamin_cost = []
    for r in range(12, 34):
        q = txt(ws.cell(row=r, column=1).value)
        if not q:
            continue
        prod = num(ws.cell(row=r, column=4).value)
        sale = num(ws.cell(row=r, column=8).value)
        aisc = num(ws.cell(row=r, column=3).value)
        if prod is not None:
            quarterly["Alphamin"][q] = prod
        alphamin_cost.append({"q": q, "aisc": aisc, "production": prod, "sales": sale})
    # Metals X：行11-31，A=季度，B=C1(A$/t)，C=AISC(A$/t)，E=产量(吨，100%口径)
    ws = wb["Metals X"]
    metalsx_cost = []
    for r in range(11, 32):
        q = txt(ws.cell(row=r, column=1).value)
        if not q:
            continue
        prod = num(ws.cell(row=r, column=5).value)
        c1 = num(ws.cell(row=r, column=2).value)
        aisc = num(ws.cell(row=r, column=3).value)
        if prod is not None:
            quarterly["Metals X"][q] = prod
        metalsx_cost.append({"q": q, "c1": c1, "aisc": aisc, "production": prod})
    # Timah：行2-22，A=季度，D=锡矿产量，E=精炼锡产量，F=精炼锡销量
    ws = wb["Timah"]
    timah_refined_q = {}
    for r in range(2, 23):
        q = txt(ws.cell(row=r, column=1).value)
        if not q:
            continue
        ore = num(ws.cell(row=r, column=4).value)
        ref = num(ws.cell(row=r, column=5).value)
        if ore is not None:
            quarterly["PT Timah"][q] = ore
        if ref is not None:
            timah_refined_q[q] = ref
    # 明苏尔季度数据已改由「季度补充」sheet 读取（见下），不再使用脚本常量。

    # --- 季度补充 sheet（行 2-8）：A=公司,B=指标,C=单位,D=数据性质,E=备注,
    #     F 列起 2023Q1..2026Q2 共 14 个季度列（None=无数据）。
    #     明苏尔矿产锡/精炼锡季度、云锡/华锡/兴业季度的唯一来源；同名期间覆盖 Sheet1 值。---
    SUPP_MAP = {
        ("云锡", "产品锡"): ("refined", "云南锡业"),
        ("云锡", "锡矿(拟合)"): ("mine", "云南锡业"),
        ("华锡有色", "锡精矿(金属吨)"): ("mine", "华锡有色"),
        ("华锡有色", "锡锭"): ("refined", "华锡有色"),
        ("兴业银锡", "矿产锡"): ("mine", "兴业银锡"),
        ("明苏尔", "矿产锡(SR+B2)"): ("mine", "明苏尔"),
        ("明苏尔", "精炼锡(Pisco)"): ("refined", "明苏尔"),
    }
    supp_data = {"mine": {}, "refined": {}}   # section -> 公司 -> {季度: 值}
    supp_meta = {}                            # (section, 公司) -> {est, est_note}
    ws = wb["季度补充"]
    supp_quarters = [txt(ws.cell(row=1, column=c).value) for c in range(6, 20)]
    for r in range(2, 9):
        comp = txt(ws.cell(row=r, column=1).value)
        ind = txt(ws.cell(row=r, column=2).value)
        target = SUPP_MAP.get((comp, ind))
        if not target:
            continue
        nature = txt(ws.cell(row=r, column=4).value) or ""
        memo = txt(ws.cell(row=r, column=5).value)
        sec, cname = target
        dd = supp_data[sec].setdefault(cname, {})
        for i, q in enumerate(supp_quarters):
            if not q:
                continue
            v = num(ws.cell(row=r, column=6 + i).value)
            if v is not None:
                dd[q] = v
        # 数据性质含估算/推算/拟合/反推/累计相减 -> 前端打 † 标记并悬停显示备注原文
        est = bool(re.search(r"估算|推算|拟合|反推|相减", nature))
        meta = supp_meta.setdefault((sec, cname), {"est": False, "est_note": None})
        meta["est"] = meta["est"] or est
        meta["est_note"] = (meta["est_note"] + "；" + memo) if meta["est_note"] and memo else (memo or meta["est_note"])

    # 合并矿表公司（Sheet1 年度/季度 <- 公司 sheet 季度 <- 季度补充表，后者覆盖）
    mine_companies = []
    for name, annual in mine_annual.items():
        data = dict(annual)
        data.update(quarterly.get(name, {}))
        data.update(supp_data["mine"].get(name, {}))
        meta = supp_meta.get(("mine", name), {})
        mine_companies.append({
            "name": name,
            "country": mine_meta[name]["country"],
            "data": data,
            "yoy": compute_yoy(data),
            "guide": mine_meta[name]["guide"],
            "guide_label": "2025 全年指引",
            "guide_progress_periods": ["2025"],
            "reason": mine_meta[name]["reason"],
            "note": mine_meta[name]["note"],
            "est": meta.get("est", False),
            "est_note": meta.get("est_note"),
        })

    # --- Sheet1 锭表（行 15-20）：J=公司,K=国家,L~O=2021~2024,Q=25指引,
    #     R/S=原因,T=2025全年,U=2026Q1 ---
    refined_companies = []
    for r in range(15, 21):
        name = txt(s1.cell(row=r, column=10).value)
        if not name:
            continue
        data = {}
        for i, year in enumerate(["2021", "2022", "2023", "2024"]):
            v = num(s1.cell(row=r, column=12 + i).value)
            if v is not None:
                data[year] = v
        v = num(s1.cell(row=r, column=20).value)  # T=2025全年
        if v is not None:
            data["2025"] = v
        v = num(s1.cell(row=r, column=21).value)  # U=2026Q1
        if v is not None:
            data["2026Q1"] = v
        # Timah 精炼季度列（Timah sheet E 列）
        if name == "PT Timah":
            data.update(timah_refined_q)
        # MSC：季度标签 20XXQ4 行的 H 列为年度实际产量（季报不披露吨数）
        if name == "MSC":
            ws = wb["马来西亚MSC（锭）"]
            for rr in range(1, ws.max_row + 1):
                a = txt(ws.cell(row=rr, column=1).value)
                h = num(ws.cell(row=rr, column=8).value)
                if a and h is not None and a.endswith("Q4"):
                    data[a[:4]] = h
        # Pisco（明苏尔）：不再用常量回填。Pisco 精炼锡季度/年度唯一来源为
        # 「季度补充」表的明苏尔/精炼锡(Pisco) 序列，并入「明苏尔」行展示，本行保留但去重。
        note = "季报不披露吨数，仅年度" if name == "MSC" else None
        if name.startswith("Pisco"):
            note = "Pisco 冶炼厂精炼锡产量并入「明苏尔」行（唯一来源：季度补充表），本行不再重复展示数据"
            data = {}
        data.update(supp_data["refined"].get(name, {}))
        meta = supp_meta.get(("refined", name), {})
        reason = txt(s1.cell(row=r, column=18).value) or txt(s1.cell(row=r, column=19).value)
        refined_companies.append({
            "name": name,
            "country": txt(s1.cell(row=r, column=11).value),
            "data": data,
            "yoy": compute_yoy(data),
            "guide": txt(s1.cell(row=r, column=17).value),  # Q=25指引
            "guide_label": "2025 全年指引",
            "guide_progress_periods": ["2025"],
            "reason": reason,
            "note": note,
            "est": meta.get("est", False),
            "est_note": meta.get("est_note"),
        })

    # --- 公司卡片成本行：Alphamin/Metals X 从 Excel 成本序列取最新值（26Q2=19,043 为指引，
    #     实际锚定 26Q1=17,968）；明苏尔/Timah/国内为文件头常量 ---
    aisc_pts = [(s["q"], s["aisc"]) for s in alphamin_cost if s.get("aisc") is not None]
    al_cost = None
    al_aisc_latest_actual = None
    if aisc_pts:
        actual_pts = [(q, v) for q, v in aisc_pts if q not in ALPHAMIN_AISC_GUIDE_PERIODS]
        q_act, v_act = actual_pts[-1]
        al_aisc_latest_actual = (q_act, v_act)
        parts = [f"AISC ${v_act:,.0f}/t（{_short_q(q_act)}）"]
        q_last, v_last = aisc_pts[-1]
        if q_last in ALPHAMIN_AISC_GUIDE_PERIODS:
            parts.append(f"{_short_q(q_last)} 指引 ${v_last:,.0f}/t（指引）")
        al_cost = "｜".join(parts)
    mx_pts = [(s["q"], s["c1"], s["aisc"]) for s in metalsx_cost
              if s.get("c1") is not None or s.get("aisc") is not None]
    mx_cost = None
    mx_aisc_latest = None
    if mx_pts:
        q_mx, c1_mx, aisc_mx = mx_pts[-1]
        mx_aisc_latest = (q_mx, aisc_mx)
        seg = []
        if c1_mx is not None:
            seg.append(f"C1 A${c1_mx:,.0f}/t")
        if aisc_mx is not None:
            seg.append(f"AISC A${aisc_mx:,.0f}/t")
        mx_cost = " · ".join(seg) + f"（{_short_q(q_mx)}）"
    mine_cost_map = {
        "Alphamin": al_cost,
        "Metals X": mx_cost,
        "明苏尔": MINSUR_COST_TEXT,
        "PT Timah": TIMAH_COST_TEXT,
        "兴业银锡": TIN_DOMESTIC_COST[("mine", "兴业银锡")],
    }
    for c in mine_companies:
        c["cost"] = mine_cost_map.get(c["name"])
    for c in refined_companies:
        c["cost"] = TIN_DOMESTIC_COST.get(("refined", c["name"]))

    # --- 锡成本对比图数据（US$/t；Metals X 按 AUDUSD_RATE 假设汇率折算，tooltip 显示澳元原值）---
    _q1, _v1 = al_aisc_latest_actual
    _qmx, _vmx = mx_aisc_latest
    cost_compare_items = [
        {"name": "Minsur 精炼锡现金成本", "value": 7751, "label": "$7,751/t",
         "note": "1Q26 实际，同比 +7%", "est": False},
        {"name": "Alphamin AISC", "value": _v1, "label": f"${_v1:,.0f}/t",
         "note": f"{_short_q(_q1)} 实际（26Q2 指引 $19,043/t）", "est": False},
        {"name": "Metals X AISC（折算）", "value": round(_vmx * AUDUSD_RATE),
         "label": f"≈${round(_vmx * AUDUSD_RATE):,}/t",
         "note": f"A${_vmx:,.0f}/t（{_short_q(_qmx)}），按 AUDUSD={AUDUSD_RATE} 假设汇率折算", "est": True},
        {"name": "Timah（券商估算）", "value": 21686, "label": "$21,686/t",
         "note": "2023 锚点，BRIDS 模型估算", "est": True},
    ]
    cost_compare_items.sort(key=lambda x: x["value"])
    cost_compare = {
        "title": "锡成本对比（US$/t，升序）",
        "items": cost_compare_items,
        "footnotes": [
            "口径差异：AISC（维持性全成本）vs 精炼现金成本 vs 券商估算不可直接横比，仅作量级参考。",
            "Alphamin 26Q1 AISC 构成（MD&A）：on-mine $9,127 + off-mine $8,056 + 维持资本 $785；其中柴油约占 $2,000/t。",
            "Minsur San Rafael 吨处理成本锚点：1Q25 $123/t、2024 全年 $143/t（US$/t treated）。",
            "Timah 官方不披露成本；BRIDS 模型敏感性 ±500$/t → 净利 ∓10-12%，极端情景 $24,000/t。",
            "国内（2025 年报）：云锡锡锭毛利率 11.45%；华锡锡锭单位成本 +26.44%、毛利率 -6.77pct；兴业整体毛利率 57.59%（-5.36pct）。",
        ],
    }

    # --- 缺季拟合（机制 1）：矿表/锭表各公司年度有值、季度有空缺时自动拟合。
    #     明苏尔锭表 2021=61/2022=55 为历史遗留异常口径，排除拟合。---
    n_fit = fit_missing_quarters(mine_companies)
    n_fit += fit_missing_quarters(refined_companies, skip_years={"明苏尔": {"2021", "2022"}})

    # 期间轴（拟合后重取，拟合可能补出早期年份的季度）
    mine_periods = sorted({p for c in mine_companies for p in c["data"] if "Q" in p}, key=quarter_sort_key)
    mine_years = sorted({p for c in mine_companies for p in c["data"] if "Q" not in p})
    ref_periods = sorted({p for c in refined_companies for p in c["data"] if "Q" in p}, key=quarter_sort_key)
    ref_years = sorted({p for c in refined_companies for p in c["data"] if "Q" not in p})

    all_periods = sorted(set(mine_periods) | set(mine_years), key=lambda p: quarter_sort_key(p) if "Q" in p else (int(p), 0))
    mine = {
        "key": "mine",
        "title": "矿山产量（矿产锡）", "unit": "吨",
        "quarters": mine_periods, "years": mine_years,
        "companies": mine_companies,
        "total": sum_total(mine_companies, mine_periods + mine_years),
    }
    refined = {
        "key": "refined",
        "title": "精炼锡产量", "unit": "吨",
        "quarters": ref_periods, "years": ref_years,
        "companies": refined_companies,
        "total": sum_total(refined_companies, ref_periods + ref_years),
    }
    return {
        "sections": [mine, refined],
        "costs": {
            "alphamin": {"title": "Alphamin AISC（US$/t）", "currency": "US$/t", "series": alphamin_cost},
            "metalsx": {"title": "Metals X C1 / AISC（A$/t，100% 口径）", "currency": "A$/t", "series": metalsx_cost},
        },
        "cost_compare": cost_compare,
        "capex": None,
        "changelog": TIN_CHANGELOG,
        "overview": section_stats(mine_companies, mine_periods, "矿山"),
        "last_update": TIN_CHANGELOG[-1]["date"],
        "_all_periods": all_periods,
        "_fitted": n_fit,
    }


# ---------------------------------------------------------------------------
# 锌抽取
# ---------------------------------------------------------------------------
def _zinc_period_map():
    """锌 sheet 列头 -> 标准期间名。'-' 表示无数据。"""
    m = {}
    for y in ("23", "24", "25", "26"):
        for q in (1, 2, 3, 4):
            m[f"{y}Q{q}"] = f"20{y}Q{q}"
        m[f"{y}总计"] = f"20{y}"
    return m


def extract_zinc(path):
    wb = load_workbook(path, data_only=False)
    pmap = _zinc_period_map()

    def read_rows(ws, rows, col_name, col_country, col_project, col_reason,
                  data_cols, yoy_cols, col_guide=None, col_cost=None):
        """通用行读取：data_cols={列号: 期间名}, yoy_cols={期间名: 列号}（Excel 已给同比，作回填）。"""
        companies = []
        for r in rows:
            name = txt(ws.cell(row=r, column=col_name).value)
            project = txt(ws.cell(row=r, column=col_project).value) if col_project else None
            if not name and not project:
                continue
            data = {}
            for col, p in data_cols.items():
                v = num(ws.cell(row=r, column=col).value)
                if v is not None:
                    data[p] = v
            provided = {}
            for p, col in yoy_cols.items():
                v = num(ws.cell(row=r, column=col).value)
                if v is not None:
                    provided[p] = v
            companies.append({
                "name": name or project,
                "project": project if (name and project and project != name) else None,
                "country": txt(ws.cell(row=r, column=col_country).value) if col_country else None,
                "data": data,
                "yoy": compute_yoy(data, provided),
                "guide": txt(ws.cell(row=r, column=col_guide).value) if col_guide else None,
                "guide_label": "FY2026 指引",
                "guide_progress_periods": ["2026Q1", "2026Q2"],
                "cost": txt(ws.cell(row=r, column=col_cost).value) if col_cost else None,
                "reason": txt(ws.cell(row=r, column=col_reason).value) if col_reason else None,
                "note": None,
            })
        return companies

    # --- 锌矿企业·季度产量：行2-22 公司，行23 总计 ---
    ws = wb["锌矿企业·季度产量"]
    mine_cols = {}
    for col in range(4, 24):  # D..W
        h = txt(ws.cell(row=1, column=col).value)
        if h in pmap:
            mine_cols[col] = pmap[h]
    mine_companies = read_rows(
        ws, range(2, 23), col_name=1, col_country=2, col_project=3, col_reason=24,
        data_cols=mine_cols,
        yoy_cols={"2025": 19, "2026Q1": 21, "2026Q2": 23},
        col_guide=25, col_cost=26,
    )
    for c in mine_companies:
        if c["name"] == "Nyrstar":
            c["note"] = "不披露季度产量财报"

    # --- 缺季拟合（机制 1）：锌矿/锌锭两表；总计行为 Excel 原值（实际披露口径），
    #     拟合仅填公司空缺季，不改总计行 ---
    n_fit = fit_missing_quarters(mine_companies)
    # 总计行（行23，B 列=总计）
    total_row = {"data": {}, "yoy": {}}
    provided = {}
    for p, col in (("2025", 19), ("2026Q1", 21), ("2026Q2", 23)):
        v = num(ws.cell(row=23, column=col).value)
        if v is not None:
            provided[p] = v
    for col, p in mine_cols.items():
        v = num(ws.cell(row=23, column=col).value)
        if v is not None:
            total_row["data"][p] = v
    total_row["yoy"] = compute_yoy(total_row["data"], provided)

    # --- 锌锭冶炼企业·季度产量：行2-13 公司，行14 总计 ---
    ws2 = wb["锌锭冶炼企业·季度产量"]
    ref_cols = {}
    for col in range(3, 23):  # C..V
        h = txt(ws2.cell(row=1, column=col).value)
        if h in pmap:
            ref_cols[col] = pmap[h]
    ref_companies = read_rows(
        ws2, range(2, 14), col_name=2, col_country=1, col_project=None, col_reason=24,
        data_cols=ref_cols,
        yoy_cols={"2024": 13, "2025": 19, "2026Q1": 21, "2026Q2": 23},
    )
    n_fit += fit_missing_quarters(ref_companies)
    ref_total = {"data": {}, "yoy": {}}
    provided = {}
    for p, col in (("2024", 13), ("2025", 19), ("2026Q1", 21), ("2026Q2", 23)):
        v = num(ws2.cell(row=14, column=col).value)
        if v is not None:
            provided[p] = v
    for col, p in ref_cols.items():
        v = num(ws2.cell(row=14, column=col).value)
        if v is not None:
            ref_total["data"][p] = v
    ref_total["yoy"] = compute_yoy(ref_total["data"], provided)

    # --- 锌 C1 成本曲线：Z 列「成本（$/t）」文本解析，$/lb 按 1$/lb=2204.62$/t 换算；
    #     Z 列为空时回退解析 FY2026 指引文本中的 'C1 $x-y/lb'（如 Kipushi）；原文保留 ---
    LB_TO_T = 2204.62
    curve_items = []
    for c in mine_companies:
        raw = c["cost"]
        lo = hi = None
        src_note = None
        if raw:
            m = re.search(r"(\d+(?:\.\d+)?)\s*[-–~]\s*(\d+(?:\.\d+)?)\s*(\$/lb)?", raw)
            if m:
                lo, hi = float(m.group(1)), float(m.group(2))
                if m.group(3):
                    lo, hi = lo * LB_TO_T, hi * LB_TO_T
                    src_note = "由 $/lb 换算"
            else:
                m1 = re.fullmatch(r"(\d+(?:\.\d+)?)", raw.strip())
                if m1:
                    lo = hi = float(m1.group(1))
        if lo is None and c["guide"]:
            m = re.search(r"C1\s*\$(\d+(?:\.\d+)?)\s*[-–~]\s*(\d+(?:\.\d+)?)\s*/lb", c["guide"], re.I)
            if m:
                lo, hi = float(m.group(1)) * LB_TO_T, float(m.group(2)) * LB_TO_T
                raw = re.search(r"C1\s*\$[^；;]*", c["guide"], re.I).group(0)
                src_note = "由指引文本 $/lb 换算"
        if lo is None:
            continue
        label = c["name"] + ("·" + c["project"] if c["project"] else "")
        curve_items.append({
            "name": label, "lo": round(lo), "hi": round(hi), "mid": round((lo + hi) / 2),
            "raw": raw, "note": src_note or "Z 列成本指引", "est": False,
        })
    curve_items.sort(key=lambda x: x["mid"])
    cost_curve = {
        "title": "锌矿 C1 成本曲线（$/t，升序）",
        "items": curve_items,
        "footnotes": [
            "单位已统一：$/lb 按 1$/lb = 2204.62$/t 换算（Antamina、Kipushi）；均为公司成本指引值（非实际），原文见表格「成本」列或条形悬停。",
            "仅含可解析出数值的公司，其余公司未给数值成本指引。",
        ],
    }

    # --- 资本开支：行2-11 原样渲染 ---
    ws3 = wb["锌矿企业·资本开支"]
    capex_headers = [txt(ws3.cell(row=1, column=c).value) for c in range(1, 10)]
    capex_rows = []
    for r in range(2, 12):
        row = []
        empty = True
        for c in range(1, 10):
            v = ws3.cell(row=r, column=c).value
            if v is not None:
                empty = False
            row.append(str(v) if v is not None else None)
        if not empty:
            capex_rows.append(row)

    # --- 更新日志：行2 起 ---
    ws4 = wb["更新日志"]
    changelog = []
    for r in range(2, ws4.max_row + 1):
        d = txt(ws4.cell(row=r, column=1).value)
        content = txt(ws4.cell(row=r, column=2).value)
        if not d and not content:
            continue
        changelog.append({"date": d, "content": content, "source": txt(ws4.cell(row=r, column=3).value)})

    mine_periods = sorted({p for c in mine_companies for p in c["data"] if "Q" in p}, key=quarter_sort_key)
    mine_years = sorted({p for c in mine_companies for p in c["data"] if "Q" not in p})
    ref_periods = sorted({p for c in ref_companies for p in c["data"] if "Q" in p}, key=quarter_sort_key)
    ref_years = sorted({p for c in ref_companies for p in c["data"] if "Q" not in p})

    mine = {
        "key": "mine",
        "title": "锌矿产量", "unit": "万金属吨",
        "quarters": mine_periods, "years": mine_years,
        "companies": mine_companies, "total": total_row,
    }
    refined = {
        "key": "refined",
        "title": "锌锭冶炼产量", "unit": "万吨精炼锌",
        "quarters": ref_periods, "years": ref_years,
        "companies": ref_companies, "total": ref_total,
    }
    return {
        "sections": [mine, refined],
        "costs": None,
        "cost_curve": cost_curve,
        "capex": {"headers": capex_headers, "rows": capex_rows},
        "changelog": changelog,
        "overview": section_stats(mine_companies, mine_periods, "锌矿"),
        "last_update": max((c["date"] or "") for c in changelog) if changelog else None,
        "_fitted": n_fit,
    }


# ---------------------------------------------------------------------------
# 铝抽取
# ---------------------------------------------------------------------------
def _read_alu_sheet(ws, capture_pending=False):
    """铝/镍产量 sheet 同构：A=公司,B=国家,C=项目/口径,D~R=23Q1..25总计,S=25同比(公式跳过),
    T=26Q1,U=26Q1同比(公式跳过),V=26Q2,W=26Q2同比(公式跳过),X=变化原因,Y=FY2026指引,Z=备注。
    末行「总计」为 SUM 公式（无缓存值），总计改为自行求和（sum_total，同比按同口径公司集合）。
    镍表末尾另有「注：...」说明行，一并跳过。
    capture_pending=True 时（镍）：季度列中的非数字文本（如「待发布(~7/27当周)」「未披露」）
    按无数据处理，但记录到 company['pending']={period: 文本}，供公司卡片展示。
    返回 (companies, total, quarters, years)。"""
    pmap = _zinc_period_map()
    data_cols = {}
    for col in range(4, 23):  # D..V
        h = txt(ws.cell(row=1, column=col).value)
        if h in pmap:
            data_cols[col] = pmap[h]
    companies = []
    for r in range(2, ws.max_row + 1):
        name = txt(ws.cell(row=r, column=1).value)
        if not name or name == "总计" or name.startswith("注"):
            continue
        data = {}
        pending = {}
        for col, p in data_cols.items():
            raw = ws.cell(row=r, column=col).value
            v = num(raw)
            if v is not None:
                data[p] = v
            elif capture_pending and isinstance(raw, str) and "Q" in p:
                s = raw.strip()
                if s and s not in ("-", "/", "—"):
                    pending[p] = s
        companies.append({
            "name": name,
            "project": txt(ws.cell(row=r, column=3).value),   # 项目/口径
            "country": txt(ws.cell(row=r, column=2).value),
            "data": data,
            "yoy": compute_yoy(data),   # 同比列是公式字符串，一律自算
            "guide": txt(ws.cell(row=r, column=25).value),    # Y=FY2026指引
            "guide_label": "FY2026 指引",
            "guide_progress_periods": [],  # 财年口径混杂（FY27/日历年），不做进度条测算
            "reason": txt(ws.cell(row=r, column=24).value),   # X=变化原因
            "note": txt(ws.cell(row=r, column=26).value),     # Z=备注（含披露频率说明）
            "pending": pending or None,
            "est": False,
            "est_note": None,
        })
    quarters = sorted({p for c in companies for p in c["data"] if "Q" in p}, key=quarter_sort_key)
    years = sorted({p for c in companies for p in c["data"] if "Q" not in p})
    n_fit = fit_missing_quarters(companies)
    # 拟合可能新增期间（如仅有年度值的年份补出季度），期间轴在拟合后重取
    quarters = sorted({p for c in companies for p in c["data"] if "Q" in p}, key=quarter_sort_key)
    years = sorted({p for c in companies for p in c["data"] if "Q" not in p})
    total = sum_total(companies, quarters + years)
    return companies, total, quarters, years, n_fit


def extract_aluminum(path):
    wb = load_workbook(path, data_only=False)
    # 板块配置：铝土矿→首板块（总览口径），氧化铝/电解铝→冶炼板块；前端按列表渲染
    sheet_specs = [
        ("bauxite", "铝土矿·季度产量", "铝土矿产量"),
        ("alumina", "氧化铝·季度产量", "氧化铝产量"),
        ("smelter", "电解铝·季度产量", "电解铝产量"),
    ]
    sections = []
    first_companies = first_quarters = None
    n_fit_total = 0
    for sec_key, sheet_name, title in sheet_specs:
        companies, total, quarters, years, n_fit = _read_alu_sheet(wb[sheet_name])
        n_fit_total += n_fit
        if first_companies is None:
            first_companies, first_quarters = companies, quarters
        sections.append({
            "key": sec_key,
            "title": title, "unit": "万吨",
            "quarters": quarters, "years": years,
            "companies": companies, "total": total,
        })
    # 更新日志：A=日期,B=更新内容,C=数据来源
    ws = wb["更新日志"]
    changelog = []
    for r in range(2, ws.max_row + 1):
        d = txt(ws.cell(row=r, column=1).value)
        content = txt(ws.cell(row=r, column=2).value)
        if not d and not content:
            continue
        changelog.append({"date": d, "content": content, "source": txt(ws.cell(row=r, column=3).value)})
    return {
        "sections": sections,
        "costs": None,
        "capex": None,
        "changelog": changelog,
        "overview": section_stats(first_companies, first_quarters, "铝土矿"),
        "last_update": max((c["date"] or "") for c in changelog) if changelog else None,
        "_fitted": n_fit_total,
    }


# ---------------------------------------------------------------------------
# 镍抽取（sheet 结构与铝同构，复用 _read_alu_sheet；单位=吨）
# ---------------------------------------------------------------------------
def extract_nickel(path):
    wb = load_workbook(path, data_only=False)
    # 板块顺序：印尼中间品排第一位（用户指定重点：边际看印尼），其后按 sheet 顺序
    sheet_specs = [
        ("intermediate", "印尼中间品·季度产量", "印尼中间品产量"),
        ("class1", "一级镍·季度产量", "一级镍产量"),
        ("npi", "镍铁NPI·季度产量", "镍铁 NPI 产量"),
    ]
    sections = []
    first_companies = first_quarters = None
    n_fit_total = 0
    for sec_key, sheet_name, title in sheet_specs:
        companies, total, quarters, years, n_fit = _read_alu_sheet(wb[sheet_name], capture_pending=True)
        n_fit_total += n_fit
        if first_companies is None:
            first_companies, first_quarters = companies, quarters
        sections.append({
            "key": sec_key,
            "title": title, "unit": "吨",
            "quarters": quarters, "years": years,
            "companies": companies, "total": total,
        })
    # 更新日志：A=日期,B=更新内容,C=数据来源
    ws = wb["更新日志"]
    changelog = []
    for r in range(2, ws.max_row + 1):
        d = txt(ws.cell(row=r, column=1).value)
        content = txt(ws.cell(row=r, column=2).value)
        if not d and not content:
            continue
        changelog.append({"date": d, "content": content, "source": txt(ws.cell(row=r, column=3).value)})
    return {
        "sections": sections,
        "costs": None,
        "capex": None,
        "changelog": changelog,
        "overview": section_stats(first_companies, first_quarters, "印尼中间品"),
        "last_update": max((c["date"] or "") for c in changelog) if changelog else None,
        "_fitted": n_fit_total,
    }


# ---------------------------------------------------------------------------
# 机制 1：缺季拟合（estimated quarters）
# ---------------------------------------------------------------------------
# 状态行识别（不拟合）：name/project 出现即排除；reason/note 需「已/完成」级措辞，
# 避免误伤「3月银漫停产」这类经营性事件描述（如兴业银锡仍在产）。
FIT_SKIP_NAME_PROJECT = ("停产", "关停", "已出售", "已售", "待补", "保养维护", "care & maintenance")
FIT_SKIP_REASON_NOTE = ("已停产", "已关停", "已出售", "完成出售", "待补", "不披露", "保养维护", "care & maintenance")


def fit_missing_quarters(companies, skip_years=None):
    """某公司某年「年度总计有值、但四个季度有空缺」-> 生成拟合季度值填入空缺：
    - 优先用该公司其他完整四季年份的季节分布（各季占比均值）分摊剩余量；
    - 没有任何完整年份时用平均（剩余量 / 空缺季数）；
    - 保证年度闭合：年度总计 = 已有季度之和 + 拟合季度之和；
    - 剩余量 <= 0（源数据季度和已超年度总计，如 Anglo 2023）-> 跳过；
    - 0 值季 = 停产，不算空缺；状态行/口径混杂行按关键词跳过；skip_years={公司: {年份}} 精确排除。
    拟合值写入 data 并打 est_q 标记（est_q_note 记录方法），同比重算。返回拟合单元格数。"""
    skip_years = skip_years or {}
    fitted_total = 0
    for c in companies:
        blob_np = " ".join(str(x) for x in (c.get("name"), c.get("project")) if x)
        blob_rn = " ".join(str(x) for x in (c.get("reason"), c.get("note")) if x)
        if any(kw in blob_np for kw in FIT_SKIP_NAME_PROJECT) or \
           any(kw in blob_rn for kw in FIT_SKIP_REASON_NOTE):
            continue
        skip = skip_years.get(c["name"], set())
        data = c["data"]
        years = sorted(int(p) for p in data if re.fullmatch(r"\d{4}", p))
        # 完整四季年份的季度占比（季节分布）
        shares = []
        for y in years:
            qs = [data.get(f"{y}Q{i}") for i in (1, 2, 3, 4)]
            if all(v is not None for v in qs) and sum(qs) > 0:
                s = sum(qs)
                shares.append([v / s for v in qs])
        mean_share = None
        if shares:
            mean_share = [sum(sh[i] for sh in shares) / len(shares) for i in range(4)]
        for y in years:
            if str(y) in skip:
                continue
            annual = data[str(y)]
            if annual is None or annual <= 0:
                continue
            qs = {i: data.get(f"{y}Q{i}") for i in (1, 2, 3, 4)}
            missing = [i for i, v in qs.items() if v is None]
            if not missing:
                continue
            known = sum(v for v in qs.values() if v is not None)
            remaining = annual - known
            if remaining <= 0:
                continue
            c.setdefault("est_q", {})
            c.setdefault("est_q_note", {})
            use_share = mean_share is not None and len(missing) < 4
            denom = sum(mean_share[i - 1] for i in missing) if use_share else 0
            for i in missing:
                if use_share and denom > 0:
                    v = remaining * mean_share[i - 1] / denom
                    note = f"拟合值（{y} 年度总计按该公司完整年份季节分布分摊）"
                else:
                    v = remaining / len(missing)
                    note = f"拟合值（{y} 年度总计平均分摊至 {len(missing)} 个空缺季）"
                data[f"{y}Q{i}"] = v   # 不四舍五入，保证年度闭合
                c["est_q"][f"{y}Q{i}"] = True
                c["est_q_note"][f"{y}Q{i}"] = note
                fitted_total += 1
        c["yoy"] = compute_yoy(data)   # 拟合后同比重算
    return fitted_total


# ---------------------------------------------------------------------------
# 机制 2：FY2026 指引解析与年化进度
# ---------------------------------------------------------------------------
def parse_guide_value(text, unit_kind):
    """把指引文本解析为品种单位的 (lo, hi)。unit_kind='吨' / '万吨' / 'kt'。
    只取文本第一段（；或;之前），避免把「2028 目标 150-160」这类远期目标误当 2026 指引；
    支持 '109-110'、'19.3-20.3万吨'、'24-29万吨'、'6,800'、'FY27：390'、'预计10左右' 等；
    年份样数字（20XX）与财年区间（2026.7–2027.6）自动跳过；明确无指引的文本返回 None。"""
    if not text:
        return None
    t = str(text).strip()
    if not t or t in ("-", "—"):
        return None
    if re.match(r"^(无|不按|不再|待补|停产)", t) or "non-guidance" in t or "撤回" in t or "暂停" in t:
        return None
    t = re.split(r"[；;]", t)[0]   # 只看第一段

    def conv(v, u):
        if unit_kind == "kt":
            if u in ("万吨", "万金属吨"):
                return v * 10
            if u == "吨":
                return v / 1000
            return v  # kt 或无单位按 kt
        if u in ("万吨", "万金属吨"):
            return v * 10000 if unit_kind == "吨" else v
        if u == "吨":
            return v if unit_kind == "吨" else v / 10000
        if u and u.lower() == "kt":
            return v * 1000 if unit_kind == "吨" else v / 10
        return v  # 无单位：按品种单位

    def to_f(tok):
        return float(tok.replace(",", ""))

    # 1) 区间（跳过财年区间：两端均在 1900-2100 且含小数点，如 2026.7–2027.6）
    for m in re.finditer(r"([\d,]+(?:\.\d+)?)\s*[-–~]\s*([\d,]+(?:\.\d+)?)\s*(万金属吨|万吨|吨|kt)?", t):
        lo, hi = to_f(m.group(1)), to_f(m.group(2))
        if 1900 <= lo <= 2100 and 1900 <= hi <= 2100 and ("." in m.group(1) or "." in m.group(2)):
            continue
        return conv(lo, m.group(3)), conv(hi, m.group(3))
    # 2) 行首裸数字（排除年份与「2026年…」）
    m = re.match(r"^([\d,]+(?:\.\d+)?)\s*(万金属吨|万吨|吨|kt)?", t)
    if m and not re.fullmatch(r"20[12]\d", m.group(1)) and not t[m.end():].startswith("年"):
        v = to_f(m.group(1))
        return conv(v, m.group(2)), conv(v, m.group(2))
    # 3) 任意位置 数字+单位
    m = re.search(r"([\d,]+(?:\.\d+)?)\s*(万金属吨|万吨|吨|kt)", t)
    if m and not (re.fullmatch(r"20[12]\d", m.group(1)) and t[m.end():].startswith("年")):
        v = to_f(m.group(1))
        return conv(v, m.group(2)), conv(v, m.group(2))
    # 4) 「：」后的数字（'FY27：72'、'2026：260'、'2026：铝商品约319'）
    m = re.search(r"[：:]\s*约?\s*([\d,]+(?:\.\d+)?)", t)
    if m:
        v = to_f(m.group(1))
        return v, v
    # 5) 兜底裸数字（跳过疑似年份）
    for m in re.finditer(r"([\d,]+(?:\.\d+)?)", t):
        if re.fullmatch(r"20[12]\d", m.group(1)):
            continue
        v = to_f(m.group(1))
        return v, v
    return None


def build_guide_progress(entry):
    """为品种生成「2026 指引 vs 年化进度」：逐板块逐公司解析指引、汇总 2026 已完成
    （26Q1 实际，如有 26Q2 则累计）、年化对比（仅 Q1 → ×4，注明季节性风险；含 Q2 → ×2）、
    状态标签（>=100% 超出 / >=90% 符合 / 否则不及；季节性名单不年化）。
    结果挂到 entry['guide_progress']，并把每股进度回写 company['guide2026'] 供卡片渲染。"""
    rows = []
    is_tin = entry["key"] == "tin"
    for sec in entry["sections"]:
        unit_kind = "万吨" if "万" in sec["unit"] else ("kt" if sec["unit"].strip().lower() == "kt" else "吨")
        for c in sec["companies"]:
            raw = c.get("guide")
            if is_tin:
                # 锡 Sheet1 T 列为 2025 指引；2026 指引用常量（仅矿山板块）
                raw = TIN_FY2026_GUIDE.get(c["name"]) if sec["key"] == "mine" else None
            parsed = parse_guide_value(raw, unit_kind) if raw else None
            lo = hi = mid = None
            if parsed:
                lo, hi = parsed
                mid = (lo + hi) / 2
            completed_periods = [p for p in ("2026Q1", "2026Q2") if c["data"].get(p) is not None]
            completed = sum(c["data"][p] for p in completed_periods) if completed_periods else None
            annualized = pct = status = note = None
            full_key = c["name"] + "·" + (c.get("project") or "")
            if c["name"] in NO_ANNUALIZE_GUIDE or full_key in NO_ANNUALIZE_GUIDE:
                status = "季节性，不年化" if c["name"] in NO_ANNUALIZE_GUIDE else "事件年，不年化"
            elif completed is not None and mid:
                factor = 2 if "2026Q2" in completed_periods else 4
                annualized = completed * factor
                pct = annualized / mid * 100
                status = "超出" if pct >= 100 else ("符合" if pct >= 90 else "不及")
                note = ("按 H1 年化，含季节性风险" if factor == 2
                        else "按 Q1 年化，含季节性风险")
            c["guide2026"] = None
            if raw and (parsed or completed is not None):
                c["guide2026"] = {
                    "raw": raw, "lo": lo, "hi": hi, "completed": completed,
                    "annualized": annualized, "pct": pct, "status": status,
                    "note": note, "unit": sec["unit"],
                }
            if raw or completed is not None:
                rows.append({
                    "section": sec["title"],
                    "name": c["name"] + (("·" + c["project"]) if c.get("project") else ""),
                    "guide_raw": raw, "lo": lo, "hi": hi,
                    "completed": completed, "annualized": annualized,
                    "pct": pct, "status": status, "note": note, "unit": sec["unit"],
                })
    # 2027 展望：按公司名匹配（允许前缀包含，兼容 'South32'/'South32·Cannington' 等写法）；
    # 无依据的公司按「线性持平」补填：优先 2026 年化、其次 2025 实际，标注平台推断。
    outlook = FY2027_OUTLOOK.get(entry["key"])
    comp_map = (outlook or {}).get("companies", {})
    data_by_name = {c["name"]: c for sec in entry["sections"] for c in sec["companies"]}
    for r in rows:
        base = r["name"].split("·")[0]
        hit = comp_map.get(r["name"]) or comp_map.get(base) or next(
            (v for k, v in comp_map.items() if base.startswith(k) or k in base or r["name"].startswith(k)), None)
        if hit:
            r["fy2027"] = hit
            continue
        cobj = data_by_name.get(base)
        anchor = r["annualized"] if r["annualized"] is not None else (cobj["data"].get("2025") if cobj else None)
        if anchor is not None:
            basis = "2026年化" if r["annualized"] is not None else "2025实际"
            r["fy2027"] = f"≈{anchor:,.0f} {r['unit']}（按{basis}线性持平外推，平台推断）"
        else:
            r["fy2027"] = "按 2026 水平持平外推（平台推断）"
    entry["outlook2027"] = {"date": outlook["date"], "total": outlook["total"]} if outlook else None
    entry["guide_progress"] = {"rows": rows}


# ---------------------------------------------------------------------------
# 2027 产量展望（手工维护，随财报节点更新；只填有依据的，来源类型随文标注：
# 官方=公司正式指引 / 计划=公司爬产或扩产计划 / 平台推断=本机构推算）
# ---------------------------------------------------------------------------
FY2027_OUTLOOK = {
    "copper": {
        "date": "2026-07-28",
        "companies": {
            "BHP": "FY27 指引大幅下调（官方，2026-07 已披露）",
            "FreePort": "Grasberg 2027 年底近满产，2027 为恢复大年（计划）",
            "Teck": "QB 继续爬坡；Anglo-Teck 合并交割后口径切换（仅剩中国审批）",
        },
        "total": "2027 矿端恢复性增长：Grasberg 近满产+QB 爬坡，对冲 BHP 下修；Cobre Panamá 若重启为上行期权（堆存矿已获批，重启时间未定）；TC 负值难快速逆转，矿紧叙事延续但边际缓和。",
    },
    "zinc": {
        "date": "2026-07-29",
        "companies": {
            "South32": "FY27 锌当量指引 20.47 万吨，与 FY26 持平（官方）",
            "Teck": "Red Dog 品位按计划继续下滑，Antamina 铜锌比波动（计划）",
            "Ivanhoe": "Kipushi 稳态运行首个完整年（计划）",
            "BOLIDEN·Garpenberg": "2027 磨矿指引 2.3Mt（3/14 地震后恢复，2026=1.5Mt）≈5-5.5 万金属吨，同比 +1.5-2 万吨恢复性增量（事件，官方指引折算）",
            "BOLIDEN·Tara": "2026 磨矿指引下调至 1.6Mt（开拓滞后），2027 看开拓赶上后的恢复性爬产（无官方指引，平台推断）",
            "Glencore": "Kazzinc 5/5 爆炸降负荷、Q2 已事故后恢复（53.7kt），2027 事故低基数上正常化（事件）；但 Lady Loretta 寿命终结（-5.1 万吨/年）+Mount Isa 资源枯竭+Antamina 锌品位下滑为结构性减量，2027 集团锌量难回 2025 年水平（官方口径）",
            "NEXA": "Cajamarquilla 5/13 火灾 Q2 影响约 7kt、6 月中已复满产，2027 冶炼端回归正常年份（事件）；矿端 2026 官方指引中枢 +6% 的增产节奏延续",
            "Polymetal·Ozernaya Minin": "新投产矿山爬坡期，2027 继续向设计产能爬升（计划）",
        },
        "total": "2027 矿端小幅恢复：Garpenberg 地震后爬产（+1.5-2 万金属吨）+Kazzinc 事故后正常化为事件性恢复增量，Dugald River/Kipushi 稳态，对冲 Red Dog 品位下滑+老矿衰竭（Lady Loretta 终结/Mount Isa 枯竭）；TC 拐点取决于新增矿山投放节奏，锭端过剩压力仍在。",
    },
    "aluminum": {
        "date": "2026-07-28",
        "companies": {
            "力拓": "AP60 2026 年底全投→2027 首个完整年（+16 万吨置换增量，计划）",
            "美铝": "San Ciprián 重启 2026 年底恢复 75%（约 17 万吨）→2027 全年贡献（计划）",
        },
        "total": "2027 海外新增兑现年：AP60 满产+越南多农一二期+San Ciprián+印度新增+印尼规划项目；中东复产是另一大增量——两座遭袭冶炼厂恢复+EGA Al Taweelah 提前复产（海湾产能占全球约 9%，6 月产量一度 -1/3）；国内 4,500 万吨天花板不变，增量全在海外，全球原铝增速约 2-3%（平台推断）；氧化铝河北新冶 740 万吨 2027Q4 投产强化原料过剩。",
    },
    "nickel": {
        "date": "2026-07-28",
        "companies": {
            "Nickel Industries": "ENC HPAL 2026/10 达产→2027 首个完整年（7.2 万吨镍/年，权益拟升 55%，计划）",
            "华友钴业": "印尼 HPAL 继续放量：华飞检修恢复+新线爬坡（计划）",
            "Eramet": "Weda Bay 2027 配额续批量是最大变量（RKAB 年度审批）",
        },
        "total": "2027 中间品继续放量（ENC 全年贡献+华友/中伟新线），NPI 受 RKAB 配额刚性约束；过剩幅度=配额政策 vs HPAL 投放的赛跑，政策底与供应压并存。",
    },
    "tin": {
        "date": "2026-07-28",
        "companies": {
            "Alphamin": "稳态 2 万吨/年，Mpama South 满产首个完整年；Bisie 深部勘探中（计划）",
            "PT Timah": "看 3 万吨目标兑现度（2026 年化 84% 不及）；陆上衰减 vs 海上采矿占比提升",
            "明苏尔": "San Rafael 品位下滑，B2 尾矿再处理项目补充（计划）",
            "兴业银锡": "银漫事故 2026 低基数（7/26 停产约1个月，选厂正常），2027 恢复性增长至事故前水平（事件）",
        },
        "total": "2027 恢复延续：Alphamin 稳产+印尼出口正常化，缅甸佤邦复产节奏是最大摆动项；若缅甸正常化+刚果新增，供应恢复斜率超预期，紧平衡转向小幅过剩。",
    },
    "lithium": {
        "date": "2026-07-28",
        "companies": {
            "Pilbara Minerals": "P1000 扩产（100 万吨/年）2027 贡献增量（计划）",
            "IGO": "Greenbushes CGP3 满产首个完整年，FY27 产量上台阶（计划）",
            "Mineral Resources": "Wodgina/Mt Marion 满产维持，Mt Marion 品位约束延续",
            "Liontown": "FY27 底 2.8Mtpa 稳态目标，2027 为地下爬坡关键年（计划）",
            "Sigma Lithium": "2025Q4 停产→2026 重启爬坡（26Q1 未披露），2027 恢复至约 24 万吨精矿年化（计划，视二期融资）",
            "Core Lithium": "2026/9 季度首批精矿→2027 复产贡献；BP33 2027 年中首矿（计划）",
            "SQM": "销量延续增长（2026 指引 +15% 以上），智利+海外扩张继续",
            "Rio Tinto": "Rincon 扩建爬坡+Sal de Vida 推进（公司管线）",
            "紫金矿业": "2026 目标 12 万吨后，2027 看拉果错二期+3Q 扩产（公司规划）",
        },
        "total": "2027 爬产继续：CGP3 满产+Goulamina 二期+Kathleen Valley 稳态+枧下窝首个完整年+麻米错达产+Finniss 回归，资源端预计再增 15-20%（平台推断）；出清逻辑让位于增量消化，价格弹性取决于储能需求斜率。",
    },
}


# ---------------------------------------------------------------------------
# 品种综述（数据驱动自动生成 + 手工观点段）
# ---------------------------------------------------------------------------
# 手工观点段：每次财报节点更新（配合年度 cron），date 必须随内容一起改。
REVIEW_COMMENTS = {
    "tin": {
        "date": "2026-07-28",
        "lines": [
            "最新财报季：Alphamin 26Q2 产锡 5,013 吨创纪录、滚动四季首达 2 万吨年化目标（AISC $17,968/t 仍处成本曲线低位）；Timah 26Q1 锡矿 6,312 吨（+96%）但年化 25,248 吨 vs 全年 3 万吨目标仅 84%，印尼陆上资源衰减是长期约束；国内云锡/华锡平稳，兴业银锡银漫矿业 7/26 安全事故致采区停产约 1 个月（选厂正常），影响 26Q3 矿产锡数百吨量级。",
            "年度看法：供给恢复是 2026 主基调（Mpama South 满产+印尼出口正常化），但缅甸佤邦复产节奏仍是最大不确定性；需求端半导体周期上行对焊料形成支撑，锡是「供给有故事、需求有β」的组合，紧平衡格局下价格弹性偏上行。",
        ],
    },
    "zinc": {
        "date": "2026-07-29",
        "lines": [
            "最新财报季：Teck Red Dog 26Q2 锌精矿 -18%（品位计划内下滑）、Antamina 权益 -62%（入选铜锌比降至 33%）；突发事件密集兑现——Garpenberg 3/14 地震致 2026 磨矿指引 -60%（26Q2 仅 0.2 万吨）、Kazzinc 5/5 爆炸降负荷（Q2 已事故后恢复 53.7kt）、Nexa Cajamarquilla 5/13 火灾（Q2 影响约 7kt，6 月中复满产）；指引表 17 家有数值指引，9 家超出、5 家不及，减量集中在老矿与事故矿。",
            "年度看法：矿端偏紧（TC 持续低位）与冶炼过剩并存，利润分配继续向矿端倾斜；2026 的事故减量（Garpenberg/Kazzinc/Nexa）在 2027 年均转为恢复性增量，叠加 Kipushi 稳态首个完整年，2027 矿端增速将高于 2026——「矿紧」叙事 2027 年边际缓和，TC 拐点与冶炼减产信号是锌价弹性的前提。",
        ],
    },
    "aluminum": {
        "date": "2026-07-28",
        "lines": [
            "最新财报季：国内电解铝贴近 4,500 万吨产能天花板运行，海外 Bell Bay（19.5 万吨）7 月永久退出、Arvida 旧线关停（AP60 置换爬坡）；中东冲突致海湾 6 月原铝产量降超 1/3；氧化铝价格回落向电解环节让利，中铝/神火/天山等电解毛利率 18-30% 处于历史高位区间。",
            "年度看法：2026 年供给刚性仍是有色中最确定的（国内天花板+海外电力约束），但 2027 年起海外新增进入投放窗口——Arvida AP60 满产（+16 万吨）、越南多农一期投产二期在建、San Ciprián 重启（约 17 万吨）、印度新增、印尼规划项目陆续兑现，供给叙事边际转松；氧化铝端河北新冶 740 万吨氢氧化铝项目 2027Q4 计划投产，原料过剩预期强化，成本让利逻辑接近尾部。",
            "α 视角：当前 18-30% 的电解毛利率是「天花板利润」而非稳态利润——2026 年铝的 α 在高利润红利属性，但 2027 年海外供应增量+氧化铝过剩双兑现前是利润兑现窗口，之后排序让位于铜锡；做多 2026 利润、警惕 2027 均值回归。",
        ],
    },
    "nickel": {
        "date": "2026-07-28",
        "lines": [
            "最新财报季：Vale 2Q26 成品镍 +4.2%（2020 以来最强 Q2）、Nornickel 25 年 19.85 万吨（-3%）；真正的主线在印尼——Weda Bay RKAB 仅批 12Mwmt（-70%）转保养、ESDM 定调不全面上调配额，政策底已现；但中间品（华友 28.7 万吨、NIC ENC HPAL 投料）仍在放量。",
            "年度看法：一级镍过剩未根本扭转，但「边际看印尼」下政策收紧给了价格地板；镍是平台内「供给最松但政策扰动最大」的品种，反弹弹性取决于 RKAB 执行力度 vs HPAL 放量速度的赛跑。",
        ],
    },
    "copper": {
        "date": "2026-07-28",
        "lines": [
            "最新财报季：FreePort 26Q2 356.5kt（Grasberg 爬产 26H2 约 65%，2027 底近满产）、BHP 25 年 1,953kt 但 FY27 指引大降、Anglo 26Q2 173.2kt；TC 深度负值确认矿端是全村最紧的环节。",
            "年度看法：矿端紧缺+冶炼负加工费是 2026 最强供给叙事，Grasberg 爬产与 Cobre Panamá 潜在重启是仅有的两个缓冲； Anglo-Teck 合并交割后口径切换需注意。供给紧张叙事强度排序第一，回调即买逻辑仍成立。",
        ],
    },
    "lithium": {
        "date": "2026-07-28",
        "lines": [
            "最新财报季（26Q1/澳矿 FY26Q3）：澳矿三强全线超指引——PLS 年化 110%、MinRes 126%、IGO 105%（Greenbushes 品位下滑但 CGP3 爬产对冲）；SQM 销量年化 93% 符合；Rio（Arcadium）87% 不及；紫金 26Q1 年化 6.5 万吨 vs 全年 12 万吨目标仅 54%，H2 需显著提速。",
            "年度看法：供给端 2026 是爬产大年——Greenbushes CGP3、Goulamina 满产、Kathleen Valley 地下爬坡、麻米错投产，资源端同比增量确定性强；成本曲线最左端（Greenbushes A$380-420/dmt）与右端（Liontown A$855-1,045）价差极大，价格若持续低迷，右端矿山（Finniss 已停产保养、Mt Cattlin 养护）退出逻辑仍在，供给出清是底部确认信号。",
            "α 视角：锂是平台内供给弹性最大的品种，过剩出清尚未走完；超额收益更可能出现在「右端成本矿山退出 + 需求端储能超预期」的右侧确认，而非左侧抄底。",
        ],
    },
}

# 跨品种供应端强弱对比（所有品种页共享，更新时改 date）
GLOBAL_SUPPLY_VIEW = {
    "date": "2026-07-28",
    "text": "2026 供应端强弱（紧→松）：铜（TC 负值、BHP 指引下修、Grasberg 受限）> 锡（恢复但缅甸不确定）≈ 锌矿（老矿减量、TC 低位）> 铝（天花板刚性但已满产，看需求）> 镍（过剩+印尼政策底）> 锂（爬产大年、出清未走完）。对应 α 排序：铜 > 锡 > 锌 > 铝 > 镍 > 锂。注意铝的动态变化：2027 年海外新增（AP60/多农/San Ciprián/印度/印尼）+ 氧化铝过剩兑现后，铝的排序大概率后移。",
}


def build_review(entry):
    """品种综述：各板块最新期合计/同比/披露进度 + 2026 指引达标统计 + 手工观点段。
    全部为数据驱动文本，front-end 直接渲染；观点段来自 REVIEW_COMMENTS（可空）。"""
    lines = []
    for sec in entry["sections"]:
        qs = sorted(sec.get("quarters", []), key=quarter_sort_key)
        latest = None
        for q in reversed(qs):
            if sec["total"]["data"].get(q) is not None:
                latest = q
                break
        if latest is None:
            continue
        lines.append({
            "section": sec["title"],
            "period": latest,
            "total": sec["total"]["data"][latest],
            "unit": sec["unit"],
            "yoy": sec["total"]["yoy"].get(latest),
            "disclosed": sum(1 for c in sec["companies"] if c["data"].get(latest) is not None),
            "n": len(sec["companies"]),
        })
    stat = {"超出": 0, "符合": 0, "不及": 0}
    n_pct = 0
    for r in entry.get("guide_progress", {}).get("rows", []):
        if r.get("pct") is not None and r.get("status") in stat:
            stat[r["status"]] += 1
            n_pct += 1
    entry["review"] = {
        "lines": lines,
        "guide": stat, "n_guide": n_pct,
        "comment": REVIEW_COMMENTS.get(entry["key"]),
        "global": GLOBAL_SUPPLY_VIEW,
    }


# ---------------------------------------------------------------------------
# 铜抽取（sheet 结构与铝/镍同构，复用 _read_alu_sheet；两板块单位不同：铜矿=kt，国内=万吨）
# ---------------------------------------------------------------------------
def _fix_bystrinsky_row(companies):
    """铜矿表 Nornickel·Bystrinsky（外贝加尔）行源数据错位：V(26Q2)=70 为年度指引量级
    （该行 2023 指引即 69-73kt/年），X(变化原因)=18.5 为季度量级（备注「26Q1 +6%」对应）。
    按口径还原：26Q1=18.5、26Q2=待发布(pending)、FY2026 指引=70、变化原因清空。
    仅在检测到错位特征时生效；源表修正后本函数自动失效。"""
    for c in companies:
        if c["name"] == "Nornickel" and (c["project"] or "").startswith("Bystrinsky"):
            if c["data"].get("2026Q2") == 70 and c.get("reason") == "18.5":
                c["data"].pop("2026Q2", None)
                c["data"].setdefault("2026Q1", 18.5)
                c["pending"] = dict(c.get("pending") or {})
                c["pending"]["2026Q2"] = "待发布(H1~7/27当周)"
                c["guide"] = "70（FY2026，Bystrinsky 精矿含铜）"
                c["reason"] = None
                c["yoy"] = compute_yoy(c["data"])


def extract_copper(path):
    wb = load_workbook(path, data_only=False)
    # 板块顺序 = sheet 顺序；单位板块级覆盖（铜矿=kt，国内=万吨）
    sheet_specs = [
        ("mine", "铜矿·季度产量", "铜矿产量（海外）", "kt"),
        ("domestic", "国内铜企·季度产量", "国内铜企产量", "万吨"),
    ]
    sections = []
    first_companies = first_quarters = None
    n_fit_total = 0
    for sec_key, sheet_name, title, unit in sheet_specs:
        companies, total, quarters, years, n_fit = _read_alu_sheet(wb[sheet_name], capture_pending=True)
        n_fit_total += n_fit
        if sec_key == "mine":
            _fix_bystrinsky_row(companies)
            # 错位修正后期间轴与总计重算（总计为 sum_total 自算，含修正后口径）
            quarters = sorted({p for c in companies for p in c["data"] if "Q" in p}, key=quarter_sort_key)
            years = sorted({p for c in companies for p in c["data"] if "Q" not in p})
            total = sum_total(companies, quarters + years)
        if first_companies is None:
            first_companies, first_quarters = companies, quarters
        sections.append({
            "key": sec_key,
            "title": title, "unit": unit,
            "quarters": quarters, "years": years,
            "companies": companies, "total": total,
        })
    # 更新日志：A=日期,B=更新内容,C=数据来源
    ws = wb["更新日志"]
    changelog = []
    for r in range(2, ws.max_row + 1):
        d = txt(ws.cell(row=r, column=1).value)
        content = txt(ws.cell(row=r, column=2).value)
        if not d and not content:
            continue
        changelog.append({"date": d, "content": content, "source": txt(ws.cell(row=r, column=3).value)})
    return {
        "sections": sections,
        "costs": None,
        "capex": None,
        "changelog": changelog,
        "overview": section_stats(first_companies, first_quarters, "铜矿"),
        "last_update": max((c["date"] or "") for c in changelog) if changelog else None,
        "_fitted": n_fit_total,
    }


# ---------------------------------------------------------------------------
# 锂抽取（sheet 结构与铜同构，复用 _read_alu_sheet；两板块：锂资源=吨LCE，锂盐冶炼=吨）
# ---------------------------------------------------------------------------
def extract_lithium(path):
    wb = load_workbook(path, data_only=False)
    sheet_specs = [
        ("resource", "锂资源·季度产量", "锂资源产量（矿山+盐湖）", "吨 LCE"),
        ("smelter", "锂盐冶炼·季度产量", "锂盐冶炼产量", "吨"),
    ]
    sections = []
    first_companies = first_quarters = None
    n_fit_total = 0
    for sec_key, sheet_name, title, unit in sheet_specs:
        companies, total, quarters, years, n_fit = _read_alu_sheet(wb[sheet_name], capture_pending=True)
        n_fit_total += n_fit
        if first_companies is None:
            first_companies, first_quarters = companies, quarters
        sections.append({
            "key": sec_key,
            "title": title, "unit": unit,
            "quarters": quarters, "years": years,
            "companies": companies, "total": total,
        })
    # 更新日志：A=日期,B=更新内容,C=数据来源
    ws = wb["更新日志"]
    changelog = []
    for r in range(2, ws.max_row + 1):
        d = txt(ws.cell(row=r, column=1).value)
        content = txt(ws.cell(row=r, column=2).value)
        if not d and not content:
            continue
        changelog.append({"date": d, "content": content, "source": txt(ws.cell(row=r, column=3).value)})
    # C1/单位现金成本（FY26 指引值，A$/dmt 精矿，来源见 research/*.json 底稿）
    cost_curve = {
        "title": "澳矿单位成本曲线（A$/dmt 精矿，FY26 指引，升序）",
        "currency": "A$/dmt",
        "items": [
            {"name": "IGO·Greenbushes", "lo": 380, "hi": 420, "mid": 400, "raw": "A$380-420/t",
             "note": "现金成本，2026-04-24 由 A$310-360 上调（品位下滑+CGP3 爬产）", "est": False},
            {"name": "Pilbara·Pilgangoora", "lo": 560, "hi": 600, "mid": 580, "raw": "A$560-600/t",
             "note": "FOB 单位运营成本指引，FY26 各季报维持", "est": False},
            {"name": "MinRes·Wodgina", "lo": 730, "hi": 800, "mid": 765, "raw": "A$730-800/dmt SC6",
             "note": "FOB 单位成本指引（SC6 口径）", "est": False},
            {"name": "MinRes·Mt Marion", "lo": 820, "hi": 890, "mid": 855, "raw": "A$820-890/dmt SC6",
             "note": "FOB 单位成本指引（SC6 口径，实际品位 3.6-4.6%）", "est": False},
            {"name": "Liontown·Kathleen Valley", "lo": 855, "hi": 1045, "mid": 950, "raw": "A$855-1,045/dmt sold",
             "note": "FY26 成本指引（sold 口径，地下矿爬坡期）", "est": False},
        ],
        "footnotes": [
            "均为 FY26（2025.7-2026.6）公司指引值（非实际成本），单位 A$/dmt 精矿，未折美元、未折 LCE。",
            "口径不一（现金成本 vs FOB 单位运营成本 vs dmt sold），横向比较仅作量级参考；Sigma/中企未披露可比成本，未列入。",
        ],
    }
    return {
        "sections": sections,
        "costs": None,
        "cost_curve": cost_curve,
        "capex": None,
        "changelog": changelog,
        "overview": section_stats(first_companies, first_quarters, "锂资源"),
        "last_update": max((c["date"] or "") for c in changelog) if changelog else None,
        "_fitted": n_fit_total,
    }


def attach_event_flags(entry, news):
    """把 news.json 中带 affects 的条目（事故/停产等产量事件）挂到对应位置：
    company=具体公司 → 公司卡片 ⚠ 行 + 指引表备注；company="*" → 品种级事件（entry['commodity_events']，
    前端在品种综述展示）。新闻已按日期降序，先命中即最新。"""
    entry["commodity_events"] = []
    for n in news:
        if n.get("commodity") != entry["name"] or not n.get("affects"):
            continue
        for aff in n["affects"]:
            cname = aff.get("company")
            note = aff.get("note") or n.get("title")
            if not cname:
                continue
            flag = {"date": n.get("date"), "note": note, "url": n.get("url")}
            if cname == "*":
                entry["commodity_events"].append(flag)
                continue
            proj = aff.get("project")
            for sec in entry["sections"]:
                for c in sec["companies"]:
                    name_hit = (c["name"] == cname or cname in c["name"])
                    proj_hit = (proj is None) or ((c.get("project") or "") == proj)
                    if name_hit and proj_hit:
                        c["event_flag"] = flag
            for r in entry.get("guide_progress", {}).get("rows", []):
                name_hit = r["name"].split("·")[0] == cname or r["name"].startswith(cname)
                proj_hit = (proj is None) or (proj in r["name"])
                if name_hit and proj_hit:
                    tag = f"⚠ {flag['date']} {note}"
                    if tag not in (r["note"] or ""):
                        r["note"] = ((r["note"] + "；" + tag) if r["note"] else tag)


# ---------------------------------------------------------------------------
# 构建
# ---------------------------------------------------------------------------
def main():
    # ---- 信息速递（data/news.json，可选；文件缺失/解析失败按空数组处理，不阻断构建）----
    news_file = BASE_DIR / "data" / "news.json"
    news = []
    if news_file.exists():
        try:
            raw_news = json.loads(news_file.read_text(encoding="utf-8"))
            news = [n for n in raw_news if isinstance(n, dict)]
            news.sort(key=lambda n: n.get("date") or "", reverse=True)  # 按日期降序，最新在前
        except Exception as e:
            print(f"[build] 警告：news.json 解析失败（{e}），按空数组处理")
            news = []
    print(f"[build] 信息速递 {len(news)} 条")

    commodities = []
    for key, cfg in COMMODITIES.items():
        print(f"[build] 抽取 {cfg['name']} <- {cfg['excel']}")
        extractor = globals()[cfg["extract"]]
        body = extractor(cfg["excel"])
        body.pop("_all_periods", None)
        n_fitted = body.pop("_fitted", 0)
        entry = {
            "key": key,
            "name": cfg["name"],
            "default_view": cfg.get("default_view", "quarter"),
            "calendar": make_calendar(cfg["calendar"], cfg["name"]),
            "caliber_notes": cfg["caliber_notes"],
            **body,
        }
        # 可选展示字段（锡/锌注册表声明了单位，供兼容保留；前端以 sections 为准）
        for opt in ("unit_mine", "unit_refined"):
            if opt in cfg:
                entry[opt] = cfg[opt]
        build_guide_progress(entry)   # 机制 2：FY2026 指引 vs 年化进度
        build_review(entry)           # 品种综述（数据驱动 + 手工观点段）
        attach_event_flags(entry, news)  # 事故/停产事件 → 公司卡片 ⚠ + 指引表备注
        commodities.append(entry)
        sec_summary = ", ".join(
            f"{s['title']} {len(s['companies'])} 家/{sum(len(c['data']) for c in s['companies'])} 数据点"
            for s in entry["sections"]
        )
        n_guide = sum(1 for r in entry["guide_progress"]["rows"] if r["lo"] is not None)
        print(f"  {cfg['name']}: {sec_summary}, 日历 {len(entry['calendar'])} 条, 日志 {len(entry['changelog'])} 条, "
              f"拟合 {n_fitted} 格, 指引表 {len(entry['guide_progress']['rows'])} 行(含数值指引 {n_guide} 家)")

    site_data = {
        "build_time": datetime.datetime.now().strftime("%Y-%m-%d %H:%M"),
        "commodities": commodities,
        "news": news,
    }
    OUT_FILE.parent.mkdir(parents=True, exist_ok=True)
    payload = json.dumps(site_data, ensure_ascii=False)
    OUT_FILE.write_text("window.SITE_DATA = " + payload + ";\n", encoding="utf-8")
    print(f"[build] 写出 {OUT_FILE} ({len(payload) // 1024} KB)")

    # ---- 缓存破防（文件名哈希）：把 app.js/style.css/data.js 复制为带内容哈希的文件名并改写
    #      index.html 引用。file:// 下浏览器可能忽略 ?v= 查询串，改文件名才能彻底杜绝旧缓存。----
    import hashlib
    index_file = BASE_DIR / "index.html"
    try:
        html = index_file.read_text(encoding="utf-8")
        for src_dir, stem, ext in ((BASE_DIR / "assets", "app", "js"),
                                   (BASE_DIR / "assets", "style", "css"),
                                   (BASE_DIR / "data", "data", "js")):
            src = src_dir / f"{stem}.{ext}"
            h = hashlib.md5(src.read_bytes()).hexdigest()[:8]
            hashed = src_dir / f"{stem}.{h}.{ext}"
            hashed.write_bytes(src.read_bytes())
            for old in src_dir.glob(f"{stem}.*.{ext}"):   # 清理旧哈希文件
                if old.name != hashed.name:
                    old.unlink()
            html = re.sub(rf"{stem}\.[\w.]*\.?{ext}(\?v=\d+)?", f"{stem}.{h}.{ext}", html)
        index_file.write_text(html, encoding="utf-8")
        print("[build] index.html 已切换到哈希文件名引用（防旧缓存）")
    except Exception as e:
        print(f"[build] 警告：哈希文件名改写失败（{e}），不影响数据")

    # ---- 抽查核对 ----
    tin = next(c for c in commodities if c["key"] == "tin")
    zinc = next(c for c in commodities if c["key"] == "zinc")
    alu = next(c for c in commodities if c["key"] == "aluminum")
    tin_sec_mine, tin_sec_ref = tin["sections"][0], tin["sections"][1]
    zinc_sec_mine = zinc["sections"][0]
    checks = []
    alphamin = next(c for c in tin_sec_mine["companies"] if c["name"] == "Alphamin")
    checks.append(("锡 Alphamin 2026Q2 产量 = 5013", alphamin["data"].get("2026Q2"), 5013))
    xyst = next(c for c in tin_sec_mine["companies"] if c["name"] == "兴业银锡")
    checks.append(("锡 Sheet1 兴业银锡 2023 = 7769", xyst["data"].get("2023"), 7769))
    dugald = next((c for c in zinc_sec_mine["companies"] if c["project"] == "Dugald River"), None)
    checks.append(("锌 Dugald River 26Q2 = 4.61", dugald["data"].get("2026Q2") if dugald else None, 4.61))
    checks.append(("锌矿总计行 26Q1 = 108.3", zinc_sec_mine["total"]["data"].get("2026Q1"), 108.3))
    # 「季度补充」sheet 接入核对
    tin_ref = {c["name"]: c for c in tin_sec_ref["companies"]}
    tin_mine = {c["name"]: c for c in tin_sec_mine["companies"]}
    checks.append(("云锡产品锡 2026Q1 = 25900", tin_ref["云南锡业"]["data"].get("2026Q1"), 25900))
    checks.append(("华锡锡精矿 2025Q3 = 1484.88", tin_mine["华锡有色"]["data"].get("2025Q3"), 1484.88))
    checks.append(("兴业矿产锡 2026Q1 = 777.33", tin_mine["兴业银锡"]["data"].get("2026Q1"), 777.33))
    checks.append(("明苏尔矿产锡 2025Q2 = 8390", tin_mine["明苏尔"]["data"].get("2025Q2"), 8390))
    checks.append(("明苏尔精炼锡 2024Q4 = 9288", tin_ref["明苏尔"]["data"].get("2024Q4"), 9288))
    yx_fit = sum(tin_mine["云南锡业"]["data"].get(f"2025Q{q}") or 0 for q in (1, 2, 3, 4))
    checks.append(("云锡锡矿(拟合) 2025 四个季度合计 = 31788", yx_fit, 31788))
    # 铝接入核对
    alu_bau, alu_alumina, alu_smelter = alu["sections"]
    alcoa_sm = next(c for c in alu_smelter["companies"] if c["name"] == "美铝 Alcoa")
    checks.append(("铝 美铝 2023Q1 电解铝 = 51.8", alcoa_sm["data"].get("2023Q1"), 51.8))
    alcoa_al = next(c for c in alu_alumina["companies"] if c["name"] == "美铝 Alcoa")
    checks.append(("铝 美铝 2023 氧化铝全年 = 1090.8", alcoa_al["data"].get("2023"), 1090.8))
    rio_b = next(c for c in alu_bau["companies"] if c["name"].startswith("力拓"))
    checks.append(("铝 力拓铝土矿 26Q2 = 1520.0", rio_b["data"].get("2026Q2"), 1520.0))
    rusal_sm = next(c for c in alu_smelter["companies"] if c["name"].startswith("俄铝"))
    checks.append(("铝 俄铝电解铝 2025 全年 = 391.8", rusal_sm["data"].get("2025"), 391.8))
    # 镍接入核对
    ni = next(c for c in commodities if c["key"] == "nickel")
    ni_inter, ni_class1, ni_npi = ni["sections"]
    norn = next(c for c in ni_class1["companies"] if c["name"] == "Nornickel")
    checks.append(("镍 Nornickel 2025 全年 = 198521", norn["data"].get("2025"), 198521))
    checks.append(("镍 Nornickel 26Q1 = 41746", norn["data"].get("2026Q1"), 41746))
    sudbury = next(c for c in ni_class1["companies"] if (c["project"] or "").startswith("Sudbury"))
    checks.append(("镍 Vale Sudbury 26Q2 = 9700", sudbury["data"].get("2026Q2"), 9700))
    antam = next(c for c in ni_npi["companies"] if c["name"] == "Antam")
    checks.append(("镍 Antam 2025 全年镍铁 = 16064", antam["data"].get("2025"), 16064))
    huayou = next(c for c in ni_inter["companies"] if c["name"] == "华友钴业")
    checks.append(("镍 华友 25 总计 = 287000", huayou["data"].get("2025"), 287000))
    # 机制 1 缺季拟合：兴业银锡 2023 四季闭合且拟合格带 est_q 标记
    xy = tin_mine["兴业银锡"]
    s23 = sum(xy["data"].get(f"2023Q{i}") or 0 for i in (1, 2, 3, 4))
    checks.append(("机制1 兴业 2023 四季合计(含拟合) = 7769", s23, 7769))
    checks.append(("机制1 兴业 2023Q2 带 est_q 标记", 1 if xy.get("est_q", {}).get("2023Q2") else 0, 1))
    # 机制 2 指引进度：PT Timah 2026 指引 30,000 吨，26Q1 6,312 → 年化 25,248（84%，不及）
    timah_row = next(r for r in tin["guide_progress"]["rows"] if r["name"] == "PT Timah")
    checks.append(("机制2 Timah 年化 = 25248", timah_row["annualized"], 25248))
    checks.append(("机制2 Timah 进度 = 84.16%", timah_row["pct"], 25248 / 30000 * 100))
    # 铜接入核对
    cu = next(c for c in commodities if c["key"] == "copper")
    cu_mine, cu_dom = cu["sections"]
    fcx = next(c for c in cu_mine["companies"] if c["name"] == "FreePort")
    checks.append(("铜 FreePort 26Q2 = 356.5", fcx["data"].get("2026Q2"), 356.5))
    vale_cu = next(c for c in cu_mine["companies"] if c["name"] == "Vale")
    checks.append(("铜 Vale 26Q1 = 102.3", vale_cu["data"].get("2026Q1"), 102.3))
    bhp_cu = next(c for c in cu_mine["companies"] if c["name"] == "BHP")
    checks.append(("铜 BHP 25 总计 = 1953", bhp_cu["data"].get("2025"), 1953))
    zijin = next(c for c in cu_dom["companies"] if c["name"] == "紫金矿业")
    checks.append(("铜 紫金 25 总计 = 109", zijin["data"].get("2025"), 109))
    anglo_cu = next(c for c in cu_mine["companies"] if c["name"] == "Anglo American")
    checks.append(("铜 Anglo 26Q2 = 173.2", anglo_cu["data"].get("2026Q2"), 173.2))
    # 锂接入核对
    li = next(c for c in commodities if c["key"] == "lithium")
    li_res, li_sm = li["sections"]
    igo = next(c for c in li_res["companies"] if c["name"] == "IGO")
    checks.append(("锂 IGO 26Q2 = 48375", igo["data"].get("2026Q2"), 48375))
    pls = next(c for c in li_res["companies"] if c["name"] == "Pilbara Minerals")
    checks.append(("锂 PLS 26Q1 = 29054.5", pls["data"].get("2026Q1"), 29054.5))
    sqm = next(c for c in li_res["companies"] if c["name"] == "SQM")
    checks.append(("锂 SQM 25 总计 = 233100", sqm["data"].get("2025"), 233100))
    shengxin = next(c for c in li_res["companies"] if c["name"] == "盛新锂能")
    checks.append(("锂 盛新资源 25 总计 = 37475", shengxin["data"].get("2025"), 37475))
    checks.append(("锂 盛新 25Q2 带 est_q 拟合标记", 1 if shengxin.get("est_q", {}).get("2025Q2") else 0, 1))
    tianqi = next(c for c in li_sm["companies"] if c["name"] == "天齐锂业")
    checks.append(("锂 天齐冶炼 25 总计 = 87900", tianqi["data"].get("2025"), 87900))
    print("[build] 抽查核对：")
    ok = True
    for label, got, want in checks:
        good = got is not None and abs(float(got) - float(want)) < 1e-6
        ok = ok and good
        print(f"  {'PASS' if good else 'FAIL'}  {label}（实际={got}）")
    print(f"[build] {'全部核对通过' if ok else '存在核对失败，请检查！'}")


if __name__ == "__main__":
    main()
